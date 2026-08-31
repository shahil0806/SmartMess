from flask import (
    Flask,
    request,
    redirect,
    url_for,
    session,
    render_template_string,
    jsonify,
    send_from_directory,
    send_file,
    Response
)

import sqlite3
import os
import secrets
import string
import base64
import hmac
import re
import json
from io import BytesIO
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import qrcode
import requests
from markupsafe import escape
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage

try:
    from pymongo import MongoClient, ASCENDING, ReturnDocument
    from pymongo.errors import DuplicateKeyError
except ImportError:  # Local SQLite mode remains available.
    MongoClient = None
    ASCENDING = 1
    ReturnDocument = None
    DuplicateKeyError = Exception

GOOGLE_SHEET_URL = os.environ.get("GOOGLE_SHEET_URL", "")
# =========================================================
# BASIC SETTINGS
# =========================================================

app = Flask(__name__)

app.secret_key = os.environ.get("SECRET_KEY", "smartmess-local-development-key")
app.config.update(
    MAX_CONTENT_LENGTH=5 * 1024 * 1024,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("COOKIE_SECURE", "0") == "1",
    PERMANENT_SESSION_LIFETIME=timedelta(minutes=15),
)

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "change-this-password")
ADMIN_RECOVERY_KEY = os.environ.get("ADMIN_RECOVERY_KEY", "").strip()
INDIA_TIMEZONE = ZoneInfo("Asia/Kolkata")
BRANCHES = [
    "AI & ML",
    "Civil (Construction Technology)",
    "Electronics (Robotics)",
    "Mechanical (CAD/CAM)",
]
HOSTEL_BLOCKS = ["BH-1", "BH-2"]

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
COLLEGE_NAME = "Government Polytechnic, Barh"
COLLEGE_LOGO = os.path.join(BASE_DIR, "college_logo.jpg")
DB_FILE = os.environ.get("DATABASE_PATH", os.path.join(BASE_DIR, "smart_mess.db"))
MONGODB_URI = os.environ.get("MONGODB_URI", "").strip()
MONGODB_DB = os.environ.get("MONGODB_DB", "smartmess").strip() or "smartmess"
USE_MONGO = MONGODB_URI.startswith("mongodb")
PHOTO_DIR = os.environ.get("PHOTO_DIR", os.path.join(BASE_DIR, "photos"))

os.makedirs(PHOTO_DIR, exist_ok=True)


# =========================================================
# DATABASE
# =========================================================

def db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


mongo_client = None
mongo_database = None
if USE_MONGO:
    if MongoClient is None:
        raise RuntimeError("pymongo is not installed")
    mongo_client = MongoClient(MONGODB_URI, serverSelectionTimeoutMS=10000)
    mongo_database = mongo_client[MONGODB_DB]


def next_mongo_id(sequence_name):
    result = mongo_database.counters.find_one_and_update(
        {"_id": sequence_name}, {"$inc": {"value": 1}}, upsert=True,
        return_document=ReturnDocument.AFTER
    )
    return result["value"]


def init_db():
    defaults = {
        "breakfast_open": "1", "breakfast_start": "07:00", "breakfast_end": "09:00",
        "lunch_open": "1", "lunch_start": "12:00", "lunch_end": "14:30",
        "dinner_open": "1", "dinner_start": "19:00", "dinner_end": "22:00",
        "breakfast_menu": "", "lunch_menu": "", "dinner_menu": "",
        "student_notice": "Welcome to SmartMess.",
        "weekly_menu": "{}",
    }

    if USE_MONGO:
        mongo_client.admin.command("ping")
        mongo_database.students.create_index([("student_uid", ASCENDING)], unique=True)
        mongo_database.students.create_index([("roll_number", ASCENDING)], unique=True)
        mongo_database.coupons.create_index([("token", ASCENDING)], unique=True)
        mongo_database.coupons.create_index([("student_uid", ASCENDING), ("meal", ASCENDING), ("generated_at", ASCENDING)])
        mongo_database.skipped_meals.create_index([("student_uid", ASCENDING), ("meal", ASCENDING), ("skip_date", ASCENDING)], unique=True)
        mongo_database.admins.create_index([("username", ASCENDING)], unique=True)
        mongo_database.complaints.create_index([("student_uid", ASCENDING), ("created_at", ASCENDING)])
        mongo_database.pin_requests.create_index([("student_uid", ASCENDING), ("status", ASCENDING)])
        mongo_database.activity_logs.create_index([("created_at", ASCENDING)])
        for key, value in defaults.items():
            mongo_database.settings.update_one({"_id": key}, {"$setOnInsert": {"value": value}}, upsert=True)
        for student in mongo_database.students.find({"$or": [{"pin_hash": {"$exists": False}}, {"pin_hash": ""}]}):
            temporary_pin = str(student["roll_number"])[-4:].zfill(4)
            mongo_database.students.update_one({"_id": student["_id"]}, {"$set": {"pin_hash": generate_password_hash(temporary_pin)}})
        return

    conn = db()
    conn.execute("""CREATE TABLE IF NOT EXISTS students (
        id INTEGER PRIMARY KEY AUTOINCREMENT, student_uid TEXT UNIQUE NOT NULL,
        name TEXT NOT NULL, roll_number TEXT UNIQUE NOT NULL, branch TEXT NOT NULL,
        hostel_room TEXT NOT NULL, photo_filename TEXT, photo_data BLOB, photo_mime TEXT,
        pin_hash TEXT, gender TEXT DEFAULT 'NOT SET', hostel_name TEXT DEFAULT 'NOT SET',
        hostel_block TEXT DEFAULT '', active INTEGER DEFAULT 1, created_at TEXT NOT NULL
    )""")
    columns = {row[1] for row in conn.execute("PRAGMA table_info(students)")}
    for column, definition in {
        "gender": "TEXT DEFAULT 'NOT SET'", "hostel_name": "TEXT DEFAULT 'NOT SET'",
        "hostel_block": "TEXT DEFAULT ''", "photo_data": "BLOB",
        "photo_mime": "TEXT", "pin_hash": "TEXT",
    }.items():
        if column not in columns:
            conn.execute(f"ALTER TABLE students ADD COLUMN {column} {definition}")
    conn.execute("""CREATE TABLE IF NOT EXISTS coupons (
        id INTEGER PRIMARY KEY AUTOINCREMENT, token TEXT UNIQUE NOT NULL,
        student_uid TEXT NOT NULL, meal TEXT NOT NULL, generated_at TEXT NOT NULL,
        expires_at TEXT NOT NULL, used_at TEXT, status TEXT NOT NULL DEFAULT 'ACTIVE'
    )""")
    conn.execute("CREATE TABLE IF NOT EXISTS settings (setting_key TEXT PRIMARY KEY, setting_value TEXT NOT NULL)")
    conn.execute("""CREATE TABLE IF NOT EXISTS skipped_meals (
        id INTEGER PRIMARY KEY AUTOINCREMENT, student_uid TEXT NOT NULL, meal TEXT NOT NULL,
        skip_date TEXT NOT NULL, created_at TEXT NOT NULL, UNIQUE(student_uid, meal, skip_date)
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS admins (
        id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL, role TEXT NOT NULL, active INTEGER DEFAULT 1,
        created_at TEXT NOT NULL
    )""")
    for table, additions in {
        "students": {"last_login":"TEXT", "pin_changed_at":"TEXT", "force_pin_change":"INTEGER DEFAULT 0"},
        "coupons": {"extension_count":"INTEGER DEFAULT 0", "extended_at":"TEXT", "extended_by":"TEXT"},
        "admins": {"last_login":"TEXT", "password_changed_at":"TEXT"},
    }.items():
        existing = {row[1] for row in conn.execute(f"PRAGMA table_info({table})")}
        for column, definition in additions.items():
            if column not in existing:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")
    conn.execute("""CREATE TABLE IF NOT EXISTS complaints (
        id INTEGER PRIMARY KEY AUTOINCREMENT, student_uid TEXT NOT NULL, meal TEXT,
        category TEXT NOT NULL, rating INTEGER, message TEXT NOT NULL, status TEXT DEFAULT 'OPEN',
        admin_reply TEXT DEFAULT '', created_at TEXT NOT NULL, updated_at TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS pin_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT, student_uid TEXT NOT NULL, hostel_room TEXT,
        reason TEXT, status TEXT DEFAULT 'PENDING', created_at TEXT NOT NULL,
        resolved_at TEXT, resolved_by TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS activity_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT, actor TEXT, role TEXT, action TEXT NOT NULL,
        details TEXT, created_at TEXT NOT NULL
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS auth_attempts (
        auth_key TEXT PRIMARY KEY, failed_attempts INTEGER DEFAULT 0, locked_until TEXT
    )""")
    for key, value in defaults.items():
        conn.execute("INSERT OR IGNORE INTO settings (setting_key, setting_value) VALUES (?, ?)", (key, value))
    for student in conn.execute("SELECT id, roll_number FROM students WHERE pin_hash IS NULL OR pin_hash = ''"):
        temporary_pin = str(student["roll_number"])[-4:].zfill(4)
        conn.execute("UPDATE students SET pin_hash = ? WHERE id = ?", (generate_password_hash(temporary_pin), student["id"]))
    for student in conn.execute("SELECT id, photo_filename FROM students WHERE photo_data IS NULL AND photo_filename IS NOT NULL"):
        path = os.path.join(PHOTO_DIR, student["photo_filename"])
        if os.path.exists(path):
            ext = os.path.splitext(path)[1].lower()
            mime = {".png": "image/png", ".webp": "image/webp"}.get(ext, "image/jpeg")
            with open(path, "rb") as photo_file:
                conn.execute("UPDATE students SET photo_data = ?, photo_mime = ? WHERE id = ?", (photo_file.read(), mime, student["id"]))

    conn.commit()
    conn.close()


init_db()


# =========================================================
# HELPERS
# =========================================================

def current_time():
    # Store India-local timestamps so the meal date remains correct on Render.
    return datetime.now(INDIA_TIMEZONE).replace(tzinfo=None)


def make_student_uid():
    return "STU-" + secrets.token_hex(5).upper()


def make_coupon_token():
    chars = string.ascii_uppercase + string.digits
    return "CPN-" + "".join(
        secrets.choice(chars) for _ in range(20)
    )


def admin_required(roles=None):
    if session.get("admin") is not True:
        return False
    if roles is None:
        return True
    return session.get("admin_role", "MAIN") in roles


def admin_scope_gender():
    role = session.get("admin_role", "MAIN")
    return "BOY" if role == "BOYS" else "GIRL" if role == "GIRLS" else ""


def row_dict(row):
    return dict(row) if row is not None else None


def get_setting(key, default=""):
    if USE_MONGO:
        item = mongo_database.settings.find_one({"_id": key})
        return item.get("value", default) if item else default
    conn = db()
    row = conn.execute("SELECT setting_value FROM settings WHERE setting_key = ?", (key,)).fetchone()
    conn.close()
    return row["setting_value"] if row else default


def save_settings(values):
    if USE_MONGO:
        for key, value in values.items():
            mongo_database.settings.update_one({"_id": key}, {"$set": {"value": value}}, upsert=True)
        return
    conn = db()
    for key, value in values.items():
        conn.execute("INSERT OR REPLACE INTO settings (setting_key, setting_value) VALUES (?, ?)", (key, value))
    conn.commit()
    conn.close()


def meal_is_available(meal):
    prefix = meal.lower()
    if get_setting(prefix + "_open", "1") != "1":
        return False, f"{meal.title()} is closed by admin."
    now = current_time().strftime("%H:%M")
    start = get_setting(prefix + "_start")
    end = get_setting(prefix + "_end")
    if start and end and not (start <= now <= end):
        return False, f"{meal.title()} coupon time is {start} to {end}."
    return True, ""


def student_by_id(student_id):
    if USE_MONGO:
        return mongo_database.students.find_one({"id": student_id})
    conn = db(); row = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone(); conn.close()
    return row_dict(row)


def student_by_roll(registration, active_only=False):
    if USE_MONGO:
        query = {"roll_number": registration}
        if active_only: query["active"] = 1
        return mongo_database.students.find_one(query)
    conn = db(); sql = "SELECT * FROM students WHERE roll_number = ?" + (" AND active = 1" if active_only else "")
    row = conn.execute(sql, (registration,)).fetchone(); conn.close(); return row_dict(row)


def student_by_uid(uid, active_only=False):
    if USE_MONGO:
        query = {"student_uid": uid}
        if active_only: query["active"] = 1
        return mongo_database.students.find_one(query)
    conn = db(); sql = "SELECT * FROM students WHERE student_uid = ?" + (" AND active = 1" if active_only else "")
    row = conn.execute(sql, (uid,)).fetchone(); conn.close(); return row_dict(row)


def count_students(extra=None):
    extra = extra or {}
    if USE_MONGO:
        return mongo_database.students.count_documents({"active": 1, **extra})
    where, params = ["active = 1"], []
    for key, value in extra.items(): where.append(f"{key} = ?"); params.append(value)
    conn = db(); count = conn.execute("SELECT COUNT(*) AS c FROM students WHERE " + " AND ".join(where), params).fetchone()["c"]; conn.close()
    return count


def add_student_record(data):
    if USE_MONGO:
        data = dict(data); data["id"] = next_mongo_id("students")
        mongo_database.students.insert_one(data); return data["id"]
    columns = list(data); placeholders = ",".join("?" for _ in columns)
    conn = db(); cursor = conn.execute(f"INSERT INTO students ({','.join(columns)}) VALUES ({placeholders})", [data[c] for c in columns]); conn.commit(); new_id = cursor.lastrowid; conn.close(); return new_id


def update_student_record(student_id, changes):
    if USE_MONGO:
        return mongo_database.students.update_one({"id": student_id}, {"$set": changes}).modified_count
    assignments = ", ".join(f"{key} = ?" for key in changes)
    conn = db(); cur = conn.execute(f"UPDATE students SET {assignments} WHERE id = ?", [*changes.values(), student_id]); conn.commit(); count = cur.rowcount; conn.close(); return count


def toggle_student_record(student_id):
    student = student_by_id(student_id)
    if student: update_student_record(student_id, {"active": 0 if student["active"] else 1})


def list_student_records(search="", gender="", hostel="", branch="", block=""):
    filters = {k: v for k, v in {"gender": gender, "hostel_name": hostel, "branch": branch, "hostel_block": block}.items() if v}
    if USE_MONGO:
        query = dict(filters)
        if search:
            safe = re.escape(search); query["$or"] = [{field: {"$regex": safe, "$options": "i"}} for field in ["name", "roll_number", "branch"]]
        return list(mongo_database.students.find(query).sort("name", ASCENDING))
    where, params = ["1=1"], []
    if search:
        where.append("(name LIKE ? OR roll_number LIKE ? OR branch LIKE ?)"); term = f"%{search}%"; params.extend([term, term, term])
    for key, value in filters.items(): where.append(f"{key} = ?"); params.append(value)
    conn = db(); rows = conn.execute("SELECT * FROM students WHERE " + " AND ".join(where) + " ORDER BY name", params).fetchall(); conn.close()
    return [dict(row) for row in rows]


def coupon_by_token(token):
    if USE_MONGO: return mongo_database.coupons.find_one({"token": token})
    conn = db(); row = conn.execute("SELECT * FROM coupons WHERE token = ?", (token,)).fetchone(); conn.close(); return row_dict(row)


def update_coupon(coupon_id, changes, required_status=None):
    if USE_MONGO:
        query = {"id": coupon_id}
        if required_status: query["status"] = required_status
        return mongo_database.coupons.update_one(query, {"$set": changes}).modified_count
    assignments = ", ".join(f"{key} = ?" for key in changes); params = list(changes.values()) + [coupon_id]
    sql = f"UPDATE coupons SET {assignments} WHERE id = ?"
    if required_status: sql += " AND status = ?"; params.append(required_status)
    conn = db(); cur = conn.execute(sql, params); conn.commit(); count = cur.rowcount; conn.close(); return count


def coupon_count(meal, date_prefix):
    if USE_MONGO:
        return mongo_database.coupons.count_documents({"meal": meal, "status": "USED", "generated_at": {"$regex": "^" + re.escape(date_prefix)}})
    conn = db(); count = conn.execute("SELECT COUNT(*) AS c FROM coupons WHERE meal=? AND status='USED' AND substr(generated_at,1,10)=?", (meal, date_prefix)).fetchone()["c"]; conn.close(); return count


def latest_coupon(uid, meal, date_prefix):
    if USE_MONGO:
        return mongo_database.coupons.find_one({"student_uid": uid, "meal": meal, "generated_at": {"$regex": "^" + re.escape(date_prefix)}}, sort=[("id", -1)])
    conn = db(); row = conn.execute("SELECT * FROM coupons WHERE student_uid=? AND meal=? AND substr(generated_at,1,10)=? ORDER BY id DESC LIMIT 1", (uid, meal, date_prefix)).fetchone(); conn.close(); return row_dict(row)


def add_coupon_record(data):
    if USE_MONGO:
        data = dict(data); data["id"] = next_mongo_id("coupons"); mongo_database.coupons.insert_one(data); return data
    columns = list(data); conn = db(); conn.execute(f"INSERT INTO coupons ({','.join(columns)}) VALUES ({','.join('?' for _ in columns)})", [data[c] for c in columns]); conn.commit(); row = conn.execute("SELECT * FROM coupons WHERE token=?", (data["token"],)).fetchone(); conn.close(); return dict(row)


def all_records():
    if USE_MONGO:
        records = list(mongo_database.coupons.find().sort("id", -1)); students = {s["student_uid"]: s for s in mongo_database.students.find()}
        for record in records:
            student = students.get(record["student_uid"], {})
            record.update({key: student.get(key) for key in ["name", "roll_number", "branch", "hostel_room"]})
        return records
    conn = db(); rows = conn.execute("""SELECT coupons.*,students.name,students.roll_number,students.branch,students.hostel_room FROM coupons LEFT JOIN students ON coupons.student_uid=students.student_uid ORDER BY coupons.id DESC""").fetchall(); conn.close(); return [dict(row) for row in rows]


def list_used_records(start_date="", end_date="", gender="", meal="", registration=""):
    records = [r for r in all_records() if r.get("status") == "USED"]
    students = {s["student_uid"]: s for s in list_student_records()}
    result = []
    for record in records:
        student = students.get(record.get("student_uid"), {})
        item = dict(record)
        item.update({key: student.get(key, item.get(key)) for key in [
            "name", "roll_number", "branch", "hostel_room", "gender", "hostel_name", "hostel_block"
        ]})
        used_date = (item.get("used_at") or item.get("generated_at") or "")[:10]
        if start_date and used_date < start_date: continue
        if end_date and used_date > end_date: continue
        if gender and item.get("gender") != gender: continue
        if meal and item.get("meal") != meal: continue
        if registration and item.get("roll_number") != registration: continue
        result.append(item)
    return result


def daily_meal_series(days, gender=""):
    end = current_time().date()
    labels, breakfast, lunch, dinner = [], [], [], []
    records = list_used_records((end - timedelta(days=days - 1)).isoformat(), end.isoformat(), gender=gender)
    for offset in range(days - 1, -1, -1):
        day = end - timedelta(days=offset)
        date_text = day.isoformat()
        labels.append(day.strftime("%d %b"))
        for meal, target in [("BREAKFAST", breakfast), ("LUNCH", lunch), ("DINNER", dinner)]:
            target.append(sum(1 for r in records if (r.get("used_at") or "")[:10] == date_text and r.get("meal") == meal))
    return {"labels": labels, "breakfast": breakfast, "lunch": lunch, "dinner": dinner}


def add_skip_record(uid, meal, skip_date):
    data = {"student_uid": uid, "meal": meal, "skip_date": skip_date,
            "created_at": current_time().strftime("%Y-%m-%d %H:%M:%S")}
    try:
        if USE_MONGO:
            data["id"] = next_mongo_id("skipped_meals"); mongo_database.skipped_meals.insert_one(data)
        else:
            conn = db(); conn.execute("INSERT INTO skipped_meals (student_uid,meal,skip_date,created_at) VALUES (?,?,?,?)",
                                      (uid, meal, skip_date, data["created_at"])); conn.commit(); conn.close()
        return True
    except (DuplicateKeyError, sqlite3.IntegrityError):
        return False


def list_skips(uid="", start_date="", end_date=""):
    if USE_MONGO:
        query = {}
        if uid: query["student_uid"] = uid
        if start_date or end_date:
            query["skip_date"] = {}
            if start_date: query["skip_date"]["$gte"] = start_date
            if end_date: query["skip_date"]["$lte"] = end_date
        return list(mongo_database.skipped_meals.find(query).sort("skip_date", -1))
    where, params = ["1=1"], []
    if uid: where.append("student_uid=?"); params.append(uid)
    if start_date: where.append("skip_date>=?"); params.append(start_date)
    if end_date: where.append("skip_date<=?"); params.append(end_date)
    conn = db(); rows = conn.execute("SELECT * FROM skipped_meals WHERE " + " AND ".join(where) + " ORDER BY skip_date DESC", params).fetchall(); conn.close()
    return [dict(row) for row in rows]


def find_admin(username):
    if USE_MONGO: return mongo_database.admins.find_one({"username": username.lower()})
    conn = db(); row = conn.execute("SELECT * FROM admins WHERE username=?", (username.lower(),)).fetchone(); conn.close(); return row_dict(row)


def list_admins():
    if USE_MONGO: return list(mongo_database.admins.find().sort("username", ASCENDING))
    conn = db(); rows = conn.execute("SELECT * FROM admins ORDER BY username").fetchall(); conn.close(); return [dict(r) for r in rows]


def add_admin_account(username, password, role):
    data = {"username": username.lower(), "password_hash": generate_password_hash(password), "role": role,
            "active": 1, "created_at": current_time().strftime("%Y-%m-%d %H:%M:%S")}
    try:
        if USE_MONGO:
            data["id"] = next_mongo_id("admins"); mongo_database.admins.insert_one(data)
        else:
            conn = db(); conn.execute("INSERT INTO admins (username,password_hash,role,active,created_at) VALUES (?,?,?,?,?)",
                                      (data["username"], data["password_hash"], role, 1, data["created_at"])); conn.commit(); conn.close()
        return True
    except (DuplicateKeyError, sqlite3.IntegrityError): return False


def log_activity(action, details="", actor=None, role=None):
    item = {"actor": actor or session.get("admin_username", "system"),
            "role": role or session.get("admin_role", "SYSTEM"), "action": action,
            "details": str(details)[:500], "created_at": current_time().strftime("%Y-%m-%d %H:%M:%S")}
    if USE_MONGO:
        item["id"] = next_mongo_id("activity_logs"); mongo_database.activity_logs.insert_one(item)
    else:
        conn=db(); conn.execute("INSERT INTO activity_logs(actor,role,action,details,created_at) VALUES(?,?,?,?,?)",
            (item["actor"],item["role"],item["action"],item["details"],item["created_at"])); conn.commit(); conn.close()


def list_activity(limit=100):
    if USE_MONGO: return list(mongo_database.activity_logs.find().sort("id",-1).limit(limit))
    conn=db(); rows=conn.execute("SELECT * FROM activity_logs ORDER BY id DESC LIMIT ?",(limit,)).fetchall(); conn.close(); return [dict(r) for r in rows]


def auth_state(key):
    if USE_MONGO: return mongo_database.auth_attempts.find_one({"_id":key}) or {"failed_attempts":0,"locked_until":""}
    conn=db(); row=conn.execute("SELECT * FROM auth_attempts WHERE auth_key=?",(key,)).fetchone(); conn.close(); return row_dict(row) or {"failed_attempts":0,"locked_until":""}


def auth_locked(key):
    state=auth_state(key); locked=state.get("locked_until") or ""
    return bool(locked and current_time() < datetime.strptime(locked,"%Y-%m-%d %H:%M:%S"))


def auth_fail(key):
    state=auth_state(key); count=int(state.get("failed_attempts",0))+1
    locked=(current_time()+timedelta(minutes=15)).strftime("%Y-%m-%d %H:%M:%S") if count>=5 else ""
    if USE_MONGO: mongo_database.auth_attempts.update_one({"_id":key},{"$set":{"failed_attempts":count,"locked_until":locked}},upsert=True)
    else:
        conn=db(); conn.execute("INSERT OR REPLACE INTO auth_attempts(auth_key,failed_attempts,locked_until) VALUES(?,?,?)",(key,count,locked)); conn.commit(); conn.close()
    return count


def auth_clear(key):
    if USE_MONGO: mongo_database.auth_attempts.delete_one({"_id":key})
    else:
        conn=db(); conn.execute("DELETE FROM auth_attempts WHERE auth_key=?",(key,)); conn.commit(); conn.close()


def main_password_ok(password):
    stored=get_setting("main_admin_password_hash","")
    return check_password_hash(stored,password) if stored else hmac.compare_digest(password,ADMIN_PASSWORD)


def delete_skip(uid, meal, skip_date):
    if USE_MONGO: return mongo_database.skipped_meals.delete_one({"student_uid":uid,"meal":meal,"skip_date":skip_date}).deleted_count
    conn=db(); cur=conn.execute("DELETE FROM skipped_meals WHERE student_uid=? AND meal=? AND skip_date=?",(uid,meal,skip_date)); conn.commit(); n=cur.rowcount; conn.close(); return n


def default_weekly_menu():
    return {day:{"BREAKFAST":"","LUNCH":"","DINNER":""} for day in ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]}


def weekly_menu():
    try:
        data=json.loads(get_setting("weekly_menu","{}")); base=default_weekly_menu()
        for day in base:
            if isinstance(data.get(day),dict): base[day].update(data[day])
        return base
    except Exception: return default_weekly_menu()


def add_complaint(uid, meal, category, rating, message):
    item={"student_uid":uid,"meal":meal,"category":category,"rating":rating,"message":message,"status":"OPEN","admin_reply":"","created_at":current_time().strftime("%Y-%m-%d %H:%M:%S"),"updated_at":""}
    if USE_MONGO: item["id"]=next_mongo_id("complaints"); mongo_database.complaints.insert_one(item)
    else:
        conn=db(); conn.execute("INSERT INTO complaints(student_uid,meal,category,rating,message,status,admin_reply,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",tuple(item[k] for k in ["student_uid","meal","category","rating","message","status","admin_reply","created_at","updated_at"])); conn.commit(); conn.close()


def list_complaints(uid=""):
    if USE_MONGO: return list(mongo_database.complaints.find({"student_uid":uid} if uid else {}).sort("id",-1))
    conn=db(); rows=conn.execute("SELECT * FROM complaints"+(" WHERE student_uid=?" if uid else "")+" ORDER BY id DESC",((uid,) if uid else ())).fetchall(); conn.close(); return [dict(r) for r in rows]


def add_pin_request(uid, room, reason):
    item={"student_uid":uid,"hostel_room":room,"reason":reason,"status":"PENDING","created_at":current_time().strftime("%Y-%m-%d %H:%M:%S"),"resolved_at":"","resolved_by":""}
    if USE_MONGO:
        if mongo_database.pin_requests.find_one({"student_uid":uid,"status":"PENDING"}): return False
        item["id"]=next_mongo_id("pin_requests"); mongo_database.pin_requests.insert_one(item)
    else:
        conn=db()
        if conn.execute("SELECT 1 FROM pin_requests WHERE student_uid=? AND status='PENDING'",(uid,)).fetchone(): conn.close(); return False
        conn.execute("INSERT INTO pin_requests(student_uid,hostel_room,reason,status,created_at,resolved_at,resolved_by) VALUES(?,?,?,?,?,?,?)",tuple(item[k] for k in ["student_uid","hostel_room","reason","status","created_at","resolved_at","resolved_by"])); conn.commit(); conn.close()
    return True


def list_pin_requests():
    if USE_MONGO: return list(mongo_database.pin_requests.find().sort("id",-1))
    conn=db(); rows=conn.execute("SELECT * FROM pin_requests ORDER BY id DESC").fetchall(); conn.close(); return [dict(r) for r in rows]


@app.after_request
def add_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    return response


def image_data_uri(filename):
    if not filename:
        return ""

    path = os.path.join(PHOTO_DIR, filename)

    if not os.path.exists(path):
        return ""

    try:
        with open(path, "rb") as f:
            encoded = base64.b64encode(
                f.read()
            ).decode("utf-8")

        ext = os.path.splitext(filename)[1].lower()

        mime = "image/jpeg"

        if ext == ".png":
            mime = "image/png"
        elif ext == ".webp":
            mime = "image/webp"

        return f"data:{mime};base64,{encoded}"

    except Exception:
        return ""


def qr_data_uri(text):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=3
    )

    qr.add_data(text)
    qr.make(fit=True)

    image = qr.make_image(
        fill_color="black",
        back_color="white"
    )

    buffer = BytesIO()
    image.save(buffer, format="PNG")

    encoded = base64.b64encode(
        buffer.getvalue()
    ).decode("utf-8")

    return "data:image/png;base64," + encoded


# =========================================================
# COMMON CSS
# =========================================================

CSS = """
<style>
* {
    box-sizing: border-box;
}

body {
    margin: 0;
    background: #f3f4f6;
    font-family: Arial, Helvetica, sans-serif;
    color: #111827;
}

.container {
    width: 94%;
    max-width: 1150px;
    margin: 30px auto;
}

.card {
    background: #ffffff;
    padding: 24px;
    border-radius: 18px;
    margin-bottom: 20px;
    box-shadow: 0 8px 25px rgba(0,0,0,.07);
}

.nav {
    display: flex;
    gap: 10px;
    flex-wrap: wrap;
    margin-bottom: 20px;
}

.btn {
    display: inline-block;
    padding: 12px 17px;
    background: #111827;
    color: white;
    text-decoration: none;
    border-radius: 10px;
    border: 0;
    cursor: pointer;
}

.green {
    background: #16a34a;
}

.blue {
    background: #2563eb;
}

.red {
    background: #dc2626;
}

.gray {
    background: #6b7280;
}

input,
select {
    width: 100%;
    padding: 13px;
    margin: 7px 0 16px;
    border: 1px solid #d1d5db;
    border-radius: 10px;
    font-size: 15px;
}

.grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
    gap: 16px;
}

.stat {
    background: white;
    padding: 20px;
    border-radius: 16px;
}

.stat h2 {
    margin: 5px 0 0;
    font-size: 30px;
}

.message {
    padding: 14px;
    border-radius: 10px;
    margin-bottom: 16px;
    background: #eff6ff;
}

.success {
    background: #dcfce7;
    color: #166534;
}

.error {
    background: #fee2e2;
    color: #991b1b;
}

.center {
    text-align: center;
}

.photo {
    width: 65px;
    height: 65px;
    border-radius: 50%;
    object-fit: cover;
}

.big-photo {
    width: 140px;
    height: 140px;
    object-fit: cover;
    border-radius: 20px;
}

.qr {
    width: 270px;
    max-width: 95%;
}

.countdown {
    font-size: 30px;
    font-weight: bold;
    margin: 15px 0;
}

table {
    width: 100%;
    border-collapse: collapse;
}

th,
td {
    padding: 11px;
    border-bottom: 1px solid #e5e7eb;
    text-align: left;
}

.badge {
    display: inline-block;
    padding: 5px 9px;
    border-radius: 999px;
    font-size: 12px;
    font-weight: bold;
}

.active-badge {
    background: #dcfce7;
    color: #166534;
}

.used-badge {
    background: #dbeafe;
    color: #1d4ed8;
}

.expired-badge {
    background: #fee2e2;
    color: #991b1b;
}

.meal-buttons {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 12px;
}

.meal-buttons button {
    padding: 16px;
    border: 0;
    border-radius: 12px;
    cursor: pointer;
    font-size: 16px;
    font-weight: bold;
    background: #111827;
    color: white;
}

@media(max-width: 650px) {
    .meal-buttons {
        grid-template-columns: 1fr;
    }

    table {
        font-size: 12px;
    }
}

/* Premium SmartMess theme */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
:root { --navy:#081b33; --blue:#2563eb; --cyan:#06b6d4; --ink:#10233f; --muted:#64748b; }
body { min-height:100vh; color:var(--ink); font-family:Inter,system-ui,-apple-system,"Segoe UI",sans-serif;
 background:radial-gradient(circle at 8% 0%,rgba(37,99,235,.16),transparent 28rem),radial-gradient(circle at 95% 8%,rgba(6,182,212,.13),transparent 25rem),#f4f7fb; }
body:before { content:""; display:block; height:5px; background:linear-gradient(90deg,var(--blue),var(--cyan)); }
.container { width:min(94%,1180px); margin:32px auto; }
h1,h2 { color:var(--navy); letter-spacing:-.035em; }
p { line-height:1.65; color:var(--muted); }
.card,.stat { background:rgba(255,255,255,.94); border:1px solid rgba(148,163,184,.18); box-shadow:0 18px 55px rgba(15,35,65,.09); }
.card { padding:clamp(22px,4vw,34px); border-radius:24px; }
.nav { gap:10px; padding:10px; background:var(--navy); border-radius:18px; box-shadow:0 16px 35px rgba(8,27,51,.18); }
.btn,.meal-buttons button { display:inline-flex; align-items:center; justify-content:center; gap:7px; padding:12px 18px; font:600 14px Inter,sans-serif;
 background:linear-gradient(135deg,#172a46,#0b1d35); border-radius:12px; transition:transform .2s,box-shadow .2s,filter .2s; }
.btn:hover,.meal-buttons button:hover { transform:translateY(-2px); box-shadow:0 10px 24px rgba(15,35,65,.2); filter:brightness(1.08); }
.green { background:linear-gradient(135deg,#16a34a,#059669); }
.blue { background:linear-gradient(135deg,#2563eb,#0891b2); }
.red { background:linear-gradient(135deg,#ef4444,#be123c); }
.gray { background:linear-gradient(135deg,#64748b,#475569); }
label { display:block; margin-top:5px; font-weight:650; font-size:14px; }
input,select { padding:14px 15px; color:var(--ink); background:#f8fafc; border:1px solid #dbe3ee; border-radius:12px; outline:none; font:500 15px Inter,sans-serif; transition:.2s; }
input:focus,select:focus { background:#fff; border-color:#60a5fa; box-shadow:0 0 0 4px rgba(37,99,235,.1); }
.grid { gap:17px; margin-bottom:22px; }
.stat { position:relative; overflow:hidden; padding:23px; border-radius:20px; color:var(--muted); }
.stat:after { content:""; position:absolute; right:-22px; bottom:-28px; width:85px; height:85px; border-radius:50%; background:rgba(37,99,235,.10); }
.stat h2 { font-size:34px; font-weight:800; color:var(--navy); }
.message { border-left:4px solid #3b82f6; }
.success { background:#ecfdf5; border-color:#22c55e; }
.error { background:#fff1f2; border-color:#f43f5e; }
.photo,.big-photo { border:4px solid #fff; box-shadow:0 8px 24px rgba(15,35,65,.16); }
.photo { border-radius:17px; }
.big-photo { border-radius:28px; }
.qr { padding:12px; border-radius:22px; background:#fff; box-shadow:0 15px 40px rgba(15,35,65,.13); }
.countdown { display:inline-block; min-width:130px; padding:9px 16px; color:#1d4ed8; background:#eff6ff; border-radius:14px; }
table { border-collapse:separate; border-spacing:0; }
th { color:#475569; background:#f8fafc; font-size:12px; text-transform:uppercase; letter-spacing:.05em; }
tr:hover td { background:#f8fbff; }
.badge { font-weight:800; }
.meal-buttons form,.meal-buttons button { width:100%; }
.meal-buttons button { min-height:74px; background:linear-gradient(145deg,#102b4f,#2563eb); }
.phase4-shell{display:grid;grid-template-columns:245px 1fr;min-height:100vh}.phase4-side{background:linear-gradient(180deg,#071a33,#102d55);padding:28px 18px;color:white;position:sticky;top:0;height:100vh}.phase4-side h2{color:white}.phase4-side a{display:block;color:#dbeafe;text-decoration:none;padding:11px 13px;margin:5px 0;border-radius:10px}.phase4-side a:hover{background:rgba(255,255,255,.12)}.phase4-main{padding:26px}.phase4-top{display:flex;justify-content:space-between;align-items:center;margin-bottom:22px}.phase4-top h1{margin:0}.phase4-metrics{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:15px}.phase4-metric{background:white;border:1px solid #e5edf7;border-radius:18px;padding:20px;box-shadow:0 10px 30px rgba(15,35,65,.07)}.phase4-metric b{font-size:30px;display:block;margin-top:8px}.install-pwa{background:#0ea5e9;color:white;border:0;border-radius:10px;padding:10px 14px;cursor:pointer}
@media(max-width:650px) { .container{margin:18px auto}.card{border-radius:20px;padding:20px}.nav{position:sticky;top:8px;z-index:20;overflow-x:auto;flex-wrap:nowrap}.nav .btn{white-space:nowrap}h1{font-size:27px} }
@media(max-width:800px){.phase4-shell{display:block}.phase4-side{position:relative;height:auto}.phase4-side nav{display:flex;overflow:auto}.phase4-side a{white-space:nowrap}.phase4-main{padding:16px}}
</style>
"""

PWA_HEAD='''<link rel="manifest" href="/manifest.json"><meta name="theme-color" content="#081b33">'''
PWA_SCRIPT='''<script>if('serviceWorker' in navigator)navigator.serviceWorker.register('/service-worker.js');let deferredPrompt;window.addEventListener('beforeinstallprompt',e=>{e.preventDefault();deferredPrompt=e;document.querySelectorAll('.install-pwa').forEach(b=>b.style.display='inline-block')});async function installSmartMess(){if(deferredPrompt){deferredPrompt.prompt();await deferredPrompt.userChoice;deferredPrompt=null}}</script>'''

DASHBOARD_CSS="""
<style>
.dash{min-height:100vh;background:#f7faff;color:#0b1838}.dash *{box-sizing:border-box}.side{position:fixed;inset:0 auto 0 0;width:250px;padding:22px 14px;background:linear-gradient(180deg,#06235c,#031b4c);color:#fff;z-index:30;overflow:auto}.brand{display:flex;gap:12px;align-items:center;padding:0 8px 18px;border-bottom:1px solid rgba(255,255,255,.14)}.brand img{width:58px;height:58px;border-radius:50%;object-fit:cover;background:#fff}.brand b{font-size:15px;line-height:1.35}.side-title{font-size:24px;font-weight:800;color:#19a7ff;padding:20px 10px}.side a{display:flex;align-items:center;gap:12px;color:#e8f1ff;text-decoration:none;padding:13px 15px;margin:5px 0;border-radius:9px;font-weight:600}.side a:hover,.side a.on{background:linear-gradient(135deg,#0568ff,#138cff);color:#fff}.side-foot{margin-top:30px;border:1px solid rgba(255,255,255,.18);padding:14px;border-radius:12px;line-height:1.7}.dash-main{margin-left:250px;min-height:100vh}.dash-top{height:72px;background:linear-gradient(90deg,#082e78,#021d55);color:#fff;display:flex;align-items:center;justify-content:space-between;padding:0 30px;position:sticky;top:0;z-index:20}.dash-top h1{font-size:23px;color:#fff;margin:0}.dash-content{padding:24px 28px}.welcome{display:flex;align-items:center;justify-content:space-between;margin-bottom:20px}.welcome h2{margin:0 0 5px;font-size:25px}.date-pill{border:1px solid #d7e2f3;background:#fff;border-radius:10px;padding:11px 16px}.metric-row{display:grid;grid-template-columns:repeat(3,1fr);gap:18px}.metric-row.small{grid-template-columns:repeat(5,1fr);margin-top:18px}.dmetric{background:#fff;border:1px solid #dce6f5;border-radius:12px;padding:18px;display:flex;gap:16px;align-items:center;box-shadow:0 4px 16px rgba(22,58,117,.05);min-width:0}.micon{width:58px;height:58px;border-radius:12px;display:grid;place-items:center;background:#e7efff;font-size:28px;flex:none}.dmetric strong{display:block;font-size:28px;margin:3px 0}.dmetric small{color:#60708d}.panel-grid{display:grid;grid-template-columns:1.45fr 1fr 1fr;gap:16px;margin-top:18px}.dpanel{background:#fff;border:1px solid #dce6f5;border-radius:12px;padding:18px;box-shadow:0 4px 16px rgba(22,58,117,.05)}.dpanel h3{margin:0 0 16px}.activity{display:flex;gap:10px;padding:10px 0;border-bottom:1px solid #edf1f7}.activity:last-child{border:0}.quick{display:grid;grid-template-columns:repeat(5,1fr);gap:16px}.quick a{border:1px solid #dbe5f4;border-radius:10px;padding:16px;text-decoration:none;color:#0b3f9e;background:#fff;font-weight:700}.quick a:first-child{background:#0753cf;color:#fff}.student-layout .metric-row{grid-template-columns:2fr 1fr}.profile-box{align-items:flex-start}.profile-box img{width:122px;height:122px;border-radius:9px;object-fit:cover;background:#e7edf5}.coupon-call{background:linear-gradient(135deg,#0753cf,#087cef);color:#fff;border-radius:12px;padding:26px;display:flex;align-items:center;gap:18px}.coupon-call .fakeqr{font-size:58px;background:#fff;padding:14px;border-radius:15px}.coupon-call h2{color:#fff}.coupon-call .btn{background:#fff;color:#0753cf}.menu-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}.menu-card{border:1px solid #dce6f5;border-radius:11px;padding:18px;min-height:150px}.menu-card:nth-child(1){background:#f1fcf5}.menu-card:nth-child(2){background:#f3f8ff}.menu-card:nth-child(3){background:#faf5ff}.student-lower{display:grid;grid-template-columns:1.25fr 1.2fr 1fr;gap:16px;margin-top:18px}.notice-bar{margin-top:16px;border:1px solid #f5b83b;background:#fffaf0;border-radius:10px;padding:13px 18px}.status-open{color:#069447;background:#dcf8e8;padding:5px 10px;border-radius:20px;font-weight:700}.status-closed{color:#596477;background:#edf0f4;padding:5px 10px;border-radius:20px;font-weight:700}.mobile-nav{display:none}
.login-page{min-height:100vh;background:#f4f8ff;display:grid;place-items:center;padding:28px}.login-shell{width:min(1120px,100%);min-height:660px;display:grid;grid-template-columns:44% 56%;background:#fff;border:1px solid #dce6f5;border-radius:24px;overflow:hidden;box-shadow:0 24px 70px rgba(5,35,92,.16)}.login-hero{position:relative;padding:38px;color:#fff;background:linear-gradient(145deg,#031d50,#063b94 67%,#0878e7);display:flex;flex-direction:column;overflow:hidden}.login-hero:after{content:"";position:absolute;width:370px;height:370px;border:70px solid rgba(11,142,255,.14);border-radius:50%;right:-260px;top:120px}.login-logo{display:flex;align-items:center;gap:14px;position:relative;z-index:1}.login-logo img{width:68px;height:68px;border-radius:50%;object-fit:cover;background:#fff;padding:3px}.login-logo strong{display:block;font-size:18px;line-height:1.3}.login-logo span{display:block;color:#bdddff;font-size:13px;margin-top:3px}.login-hero-content{position:relative;z-index:1;margin:auto 0}.login-hero-content h1{color:#fff;font-size:40px;margin:0 0 10px}.login-hero-content>p{color:#cfe3ff;font-size:17px}.login-features{display:grid;gap:14px;margin-top:30px}.login-feature{display:flex;align-items:center;gap:13px;font-weight:650}.login-feature i{font-style:normal;width:38px;height:38px;display:grid;place-items:center;border:1px solid #1598ff;border-radius:10px;color:#43c8ff}.login-building{font-size:12px;letter-spacing:3px;color:#8fc9ff;opacity:.7;margin-top:auto;position:relative;z-index:1}.login-pane{padding:50px;display:flex;align-items:center;background:radial-gradient(circle at 100% 0,#eaf4ff 0,transparent 38%),#fbfdff}.login-box{width:100%;max-width:450px;margin:auto}.login-box h2{text-align:center;font-size:31px;margin:0 0 7px}.login-box .sub{text-align:center;color:#667590;margin:0 0 24px}.login-tabs{display:grid;grid-template-columns:1fr 1fr;background:#f1f5fb;border:1px solid #dce4f0;border-radius:10px;overflow:hidden;margin-bottom:24px}.login-tabs a{text-align:center;padding:13px;color:#344158;text-decoration:none;font-weight:800}.login-tabs a.on{color:#fff;background:linear-gradient(135deg,#0760e9,#0a86f5)}.field-wrap{position:relative}.field-wrap input{padding-right:52px}.eye-btn{position:absolute;right:7px;top:7px;width:42px;height:42px;border:0;background:transparent;cursor:pointer;font-size:18px;border-radius:8px}.eye-btn:hover{background:#eaf2ff}.login-primary{width:100%;border:0;border-radius:10px;background:linear-gradient(135deg,#0753cf,#078bf0);color:#fff;padding:15px;font-size:16px;font-weight:800;cursor:pointer;box-shadow:0 10px 22px rgba(7,83,207,.2)}.login-primary:hover{filter:brightness(1.06)}.login-help{text-align:center;margin-top:17px}.login-help a{font-weight:700;color:#0753cf}.security-note{display:flex;gap:11px;align-items:center;margin-top:25px;padding:15px;border-top:1px solid #e1e8f2;color:#60708b;font-size:13px;line-height:1.45}.security-note b{font-size:25px;color:#0872e5}
.login-feature i svg,.tab-icon svg,.security-icon svg{width:21px;height:21px;display:block;stroke:currentColor;fill:none;stroke-width:1.9;stroke-linecap:round;stroke-linejoin:round}.login-tabs a{display:flex;align-items:center;justify-content:center;gap:8px}.tab-icon{display:inline-grid;place-items:center}.tab-icon svg{width:19px;height:19px}.admin-login-box{max-width:470px;background:#fff;border:1px solid #dce7f6;border-radius:22px;padding:28px 30px;box-shadow:0 18px 48px rgba(5,48,112,.12);position:relative;overflow:hidden}.admin-login-box:before{content:"";position:absolute;inset:0 0 auto;height:5px;background:linear-gradient(90deg,#075be0,#09a3ee)}.admin-login-box h2{font-size:28px}.role-label{display:block;font-size:13px;font-weight:800;color:#263650;margin:0 0 9px}.role-options{display:grid;grid-template-columns:repeat(3,1fr);gap:7px;margin-bottom:19px}.role-options input{position:absolute;opacity:0;pointer-events:none}.role-options label{margin:0;padding:10px 5px;text-align:center;border:1px solid #d9e3f1;border-radius:9px;background:#f6f9fd;color:#53637c;font-size:12px;font-weight:800;cursor:pointer}.role-options input:checked+label{color:#0758d4;background:#eaf3ff;border-color:#3487ec;box-shadow:0 0 0 2px rgba(52,135,236,.1)}.remember-help{display:flex;align-items:center;justify-content:space-between;gap:12px;margin:12px 0 18px;font-size:13px}.remember-check{display:flex;align-items:center;gap:8px;color:#53637c}.remember-check input{width:16px;height:16px;margin:0}.remember-help a{font-weight:750;color:#0753cf}.security-icon{display:grid;place-items:center;color:#0872e5}.security-icon svg{width:24px;height:24px}
@media(max-width:1050px){.metric-row.small{grid-template-columns:repeat(3,1fr)}.panel-grid,.student-lower{grid-template-columns:1fr}.quick{grid-template-columns:repeat(2,1fr)}}
@media(max-width:760px){.side{display:none}.dash-main{margin-left:0}.dash-top{padding:0 15px}.dash-top h1{font-size:18px}.dash-content{padding:15px}.metric-row,.metric-row.small,.student-layout .metric-row,.menu-grid{grid-template-columns:1fr}.welcome{align-items:flex-start;gap:12px}.date-pill{font-size:12px}.mobile-nav{display:block}.profile-box{flex-direction:column}.quick{grid-template-columns:1fr}.login-page{padding:12px}.login-shell{grid-template-columns:1fr;border-radius:18px}.login-hero{padding:22px;min-height:210px}.login-logo img{width:52px;height:52px}.login-logo strong{font-size:15px}.login-hero-content{margin:24px 0 0}.login-hero-content h1{font-size:28px}.login-hero-content>p{font-size:14px}.login-features,.login-building{display:none}.login-pane{padding:28px 22px}.login-box h2{font-size:26px}.admin-login-box{padding:26px 20px}.role-options{grid-template-columns:1fr}.remember-help{align-items:flex-start;flex-direction:column}}
</style>
"""


# =========================================================
# ROOT
# =========================================================

@app.route("/")
def root():
    return redirect(url_for("student_home"))


@app.route("/health")
def health():
    return jsonify({
        "status": "ok",
        "service": "SmartMess",
        "database": "mongodb" if USE_MONGO else "sqlite",
    })


# =========================================================
# ADMIN LOGIN
# =========================================================

@app.route("/admin", methods=["GET", "POST"])
def admin_login():

    if admin_required():
        return redirect(url_for("admin_dashboard"))

    error = ""

    if request.method == "POST":

        username = request.form.get("username", "main").strip().lower() or "main"
        password = request.form.get("password", "")

        key="admin:"+username
        if auth_locked(key):
            error="Too many wrong attempts. Try again after 15 minutes."
        elif username in ["main", "admin"] and main_password_ok(password):

            auth_clear(key)
            session["admin"] = True
            session["admin_role"] = "MAIN"
            session["admin_username"] = "main"
            session.permanent=True
            log_activity("ADMIN_LOGIN","Main admin login",actor="main",role="MAIN")

            return redirect(url_for("admin_dashboard"))

        account = find_admin(username)
        if not error and account and account.get("active", 1) and check_password_hash(account["password_hash"], password):
            auth_clear(key)
            session["admin"] = True
            session["admin_role"] = account["role"]
            session["admin_username"] = account["username"]
            session.permanent=True
            now=current_time().strftime("%Y-%m-%d %H:%M:%S")
            if USE_MONGO: mongo_database.admins.update_one({"id":account["id"]},{"$set":{"last_login":now}})
            else:
                conn=db(); conn.execute("UPDATE admins SET last_login=? WHERE id=?",(now,account["id"])); conn.commit(); conn.close()
            log_activity("ADMIN_LOGIN","Successful login")
            return redirect(url_for("admin_scanner" if account["role"] == "SCANNER" else "admin_dashboard"))

        if not error:
            attempts=auth_fail(key); error = "Wrong username or password."
            log_activity("ADMIN_LOGIN_FAILED",f"Username: {username}; attempt {attempts}",actor=username,role="UNKNOWN")

    html = f"""<!doctype html><html><head><title>Admin Login | SmartMess</title>{CSS}{DASHBOARD_CSS}{PWA_HEAD}</head><body class="login-page">
    <main class="login-shell"><section class="login-hero"><div class="login-logo"><img src="/static/icon-192.png" alt="College logo"><div><strong>Government Polytechnic, Barh</strong><span>SmartMess Management System</span></div></div><div class="login-hero-content"><h1>SmartMess</h1><p>Smart Mess Management System</p><div class="login-features"><div class="login-feature"><i><svg viewBox="0 0 24 24"><path d="M16 21v-2a4 4 0 0 0-4-4H6a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M22 21v-2a4 4 0 0 0-3-3.87M16 3.13a4 4 0 0 1 0 7.75"/></svg></i> Manage Students</div><div class="login-feature"><i><svg viewBox="0 0 24 24"><path d="M3 8V5a2 2 0 0 1 2-2h3M16 3h3a2 2 0 0 1 2 2v3M21 16v3a2 2 0 0 1-2 2h-3M8 21H5a2 2 0 0 1-2-2v-3"/><rect x="8" y="8" width="3" height="3"/><rect x="13" y="8" width="3" height="3"/><rect x="8" y="13" width="3" height="3"/><path d="M14 14h2v2"/></svg></i> QR Scanner &amp; Approval</div><div class="login-feature"><i><svg viewBox="0 0 24 24"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><path d="M14 2v6h6M8 13h8M8 17h8M8 9h2"/></svg></i> Reports &amp; Activity Log</div></div></div><div class="login-building">GOVERNMENT POLYTECHNIC · BARH</div></section>
    <section class="login-pane"><div class="login-box admin-login-box"><nav class="login-tabs"><a href="/student"><span class="tab-icon"><svg viewBox="0 0 24 24"><path d="m3 10 9-5 9 5-9 5zM7 12v5c3 2 7 2 10 0v-5M21 10v6"/></svg></span>Student</a><a class="on" href="/admin"><span class="tab-icon"><svg viewBox="0 0 24 24"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/><circle cx="12" cy="9" r="2.2"/><path d="M8.5 16c.7-2 2-3 3.5-3s2.8 1 3.5 3"/></svg></span>Admin</a></nav><h2>Admin Portal</h2><p class="sub">Authorized staff access only</p>{'<div class="message error">' + escape(error) + '</div>' if error else ''}<form method="post"><span class="role-label">Login as</span><div class="role-options"><input id="roleMain" type="radio" name="login_role" value="MAIN" checked><label for="roleMain">Main Admin</label><input id="roleHostel" type="radio" name="login_role" value="HOSTEL"><label for="roleHostel">Hostel Admin</label><input id="roleScanner" type="radio" name="login_role" value="SCANNER"><label for="roleScanner">Scanner Operator</label></div><label>Username</label><input name="username" value="main" autocomplete="username" required><label>Password</label><div class="field-wrap"><input id="loginSecret" type="password" name="password" autocomplete="current-password" required><button type="button" class="eye-btn" onclick="toggleSecret()" aria-label="Show password"><svg viewBox="0 0 24 24" width="21" height="21" fill="none" stroke="currentColor" stroke-width="2"><path d="M2 12s3.5-7 10-7 10 7 10 7-3.5 7-10 7S2 12 2 12z"/><circle cx="12" cy="12" r="3"/></svg></button></div><div class="remember-help"><label class="remember-check"><input type="checkbox" name="remember"> Remember this device</label><a href="/admin/forgot-password">Forgot Password?</a></div><button class="login-primary" type="submit">Secure Admin Login</button></form><div class="security-note"><span class="security-icon"><svg viewBox="0 0 24 24"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/><path d="m9 12 2 2 4-4"/></svg></span><span>All admin activity is securely monitored.</span></div></div></section></main>
    <script>function toggleSecret(){{const x=document.getElementById('loginSecret');x.type=x.type==='password'?'text':'password'}};</script>{PWA_SCRIPT}</body></html>"""

    return html


@app.route("/admin/forgot-password",methods=["GET","POST"])
def admin_forgot_password():
    message=""; ok=False; key="recovery:"+(request.remote_addr or "unknown")
    if request.method=="POST":
        username=request.form.get("username","").strip().lower(); recovery=request.form.get("recovery_key",""); password=request.form.get("new_password","")
        if auth_locked(key): message="Reset is locked for 15 minutes."
        elif not ADMIN_RECOVERY_KEY: message="Recovery is not configured. Set ADMIN_RECOVERY_KEY on Render."
        elif username not in ["main","admin"] or not hmac.compare_digest(recovery,ADMIN_RECOVERY_KEY):
            auth_fail(key); message="Username or Recovery Key is incorrect."
        elif len(password)<8: message="New password must contain at least 8 characters."
        else:
            save_settings({"main_admin_password_hash":generate_password_hash(password)})
            auth_clear(key); ok=True; message="Password changed. You can now login."
            log_activity("MAIN_PASSWORD_RESET","Recovery key reset",actor="main",role="MAIN")
    return f'''<!doctype html><html><head><title>Admin Password Recovery</title>{CSS}</head><body><div class="container"><div class="card" style="max-width:540px;margin:70px auto"><h1>🔑 Admin Password Recovery</h1><p>Only Main Admin can use the private recovery key.</p>{f'<div class="message {"success" if ok else "error"}">{escape(message)}</div>' if message else ''}<form method="post"><label>Username</label><input name="username" value="main" required><label>Private Recovery Key</label><input type="password" name="recovery_key" required><label>New Password</label><input type="password" name="new_password" minlength="8" required><button class="btn green">Reset Password</button> <a class="btn gray" href="/admin">Back</a></form></div></div></body></html>'''


@app.route("/admin/logout")
def admin_logout():
    session.clear()

    return redirect(url_for("admin_login"))

# =========================================================
# ADMIN VERIFY COUPON
# =========================================================

@app.route("/admin/verify-coupon", methods=["POST"])
def verify_coupon():

    if not admin_required():
        return jsonify({
            "success": False,
            "message": "Admin login required."
        }), 401

    data = request.get_json(silent=True) or {}

    token = str(
        data.get("token", "")
    ).strip()

    if not token:
        return jsonify({
            "success": False,
            "message": "Invalid QR."
        }), 400

    coupon = coupon_by_token(token)

    if not coupon:
        return jsonify({
            "success": False,
            "message": "Coupon not found."
        }), 404

    # ---------------------------------------------------------
    # CHECK ALREADY USED
    # ---------------------------------------------------------

    if coupon["status"] == "USED":

        return jsonify({
            "success": False,
            "message": "Coupon already used."
        })

    # ---------------------------------------------------------
    # CHECK EXPIRY
    # ---------------------------------------------------------

    expiry = datetime.strptime(
        coupon["expires_at"],
        "%Y-%m-%d %H:%M:%S"
    )

    if current_time() > expiry:

        update_coupon(coupon["id"], {"status": "EXPIRED"})

        return jsonify({
            "success": False,
            "message": "Coupon expired. 5-minute validity ended.",
            "can_extend": (coupon.get("generated_at") or "")[:10] == current_time().date().isoformat() and int(coupon.get("extension_count", 0) or 0) < 1,
            "token": token
        })

    # ---------------------------------------------------------
    # FIND STUDENT
    # ---------------------------------------------------------

    student = student_by_uid(coupon["student_uid"], active_only=True)

    if not student:

        return jsonify({
            "success": False,
            "message": "Student is not active."
        })

    # ---------------------------------------------------------
    # MARK COUPON AS USED
    # ---------------------------------------------------------

    used_time = current_time().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    if update_coupon(coupon["id"], {"status": "USED", "used_at": used_time}, required_status="ACTIVE") != 1:

        return jsonify({
            "success": False,
            "message": "Coupon was already used."
        })

    # ---------------------------------------------------------
    # SAVE SUCCESSFUL SCAN TO GOOGLE SHEET
    # ---------------------------------------------------------
    try:

        if not GOOGLE_SHEET_URL:
            raise ValueError("GOOGLE_SHEET_URL is not configured")

        now = current_time()

        response = requests.post(
            GOOGLE_SHEET_URL,
            json={
                "date": now.strftime("%d-%m-%Y"),
                "time": now.strftime("%I:%M:%S %p"),
                "name": student["name"],
                "roll": student["roll_number"],
                "branch": student["branch"],
                "room": student["hostel_room"],
                "gender": student["gender"],
                "hostel": student["hostel_name"],
                "block": student["hostel_block"],
                "meal": coupon["meal"],
                "status": "USED"
            },
            timeout=10
        )

        print(
            "GOOGLE SHEET:",
            response.status_code,
            response.text
        )

    except Exception as e:

        print(
            "Google Sheet Error:",
            e
        )
    # ---------------------------------------------------------
    # STUDENT PHOTO
    # ---------------------------------------------------------

    photo_url = ""

    if student.get("photo_filename") or student.get("photo_data"):

        photo_url = (
            "/student-photo/"
            + str(student["id"])
        )

    # ---------------------------------------------------------
    # SUCCESS RESPONSE
    # ---------------------------------------------------------

    return jsonify({
        "success": True,
        "name": student["name"],
        "roll": student["roll_number"],
        "branch": student["branch"],
        "room": student["hostel_room"],
        "gender": student["gender"],
        "hostel": student["hostel_name"],
        "block": student["hostel_block"],
        "meal": coupon["meal"],
        "photo": photo_url
    })# =========================================================
# ADMIN DASHBOARD
# =========================================================

@app.route("/admin/dashboard")
def admin_dashboard():

    if not admin_required():
        return redirect(url_for("admin_login"))
    if session.get("admin_role") == "SCANNER":
        return redirect(url_for("admin_scanner"))

    today = current_time().strftime("%Y-%m-%d")

    scope_gender = admin_scope_gender()
    scope = {"gender": scope_gender} if scope_gender else {}
    student_count = count_students(scope)
    boys_count = count_students({"gender": "BOY"}) if not scope_gender or scope_gender == "BOY" else 0
    girls_count = count_students({"gender": "GIRL"}) if not scope_gender or scope_gender == "GIRL" else 0
    bh1_count = count_students({**scope, "hostel_block": "BH-1"})
    bh2_count = count_students({**scope, "hostel_block": "BH-2"})
    today_records = list_used_records(today, today, gender=scope_gender)
    breakfast = sum(r["meal"] == "BREAKFAST" for r in today_records)
    lunch = sum(r["meal"] == "LUNCH" for r in today_records)
    dinner = sum(r["meal"] == "DINNER" for r in today_records)
    present_uids = {r["student_uid"] for r in today_records}
    absent = max(student_count - len(present_uids), 0)
    skipped = len(list_skips(start_date=today, end_date=today))
    chart7 = json.dumps(daily_meal_series(7, scope_gender))
    database_label = "MongoDB Atlas (Permanent)" if USE_MONGO else "SQLite (Local)"
    total_meals = breakfast + lunch + dinner
    recent_html = "".join(
        f'<div class="activity"><span>✅</span><div><b>{escape(item.get("action", "Activity").replace("_", " ").title())}</b><br><small>{escape(item.get("details", "") or "SmartMess update")} · {escape(item.get("created_at", ""))}</small></div></div>'
        for item in list_activity(5)
    ) or '<div class="activity"><span>ℹ️</span><div><b>No recent activity</b><br><small>New activity will appear here.</small></div></div>'
    role_label = {"MAIN":"Super Administrator","BOYS":"Boys Hostel Admin","GIRLS":"Girls Hostel Admin"}.get(session.get("admin_role"),"Administrator")
    date_label = current_time().strftime("%d %b %Y, %A")

    html = f"""<!doctype html><html><head><title>Admin Dashboard</title>{CSS}{DASHBOARD_CSS}{PWA_HEAD}<script src="https://cdn.jsdelivr.net/npm/chart.js"></script></head><body class="dash">
    <aside class="side"><div class="brand"><img src="/static/icon-192.png"><b>{COLLEGE_NAME}</b></div><div class="side-title">🍽️ SmartMess</div><nav>
      <a class="on" href="/admin/dashboard">▦ Admin Dashboard</a><a href="/admin/scanner">▣ Scanner</a><a href="/admin/students">♙ Students</a><a href="/admin/meal-settings">♨ Meal Settings</a><a href="/admin/weekly-menu">▦ Weekly Menu</a><a href="/admin/reports">▤ Reports</a><a href="/admin/complaints">◉ Complaints</a>{'<a href="/admin/activity-log">▣ Activity Log</a><a href="/admin/roles">⚙ Admin Roles</a>' if admin_required(['MAIN']) else ''}<a href="/admin/logout">↪ Logout</a></nav><div class="side-foot">📅 Today<br><b>{date_label}</b></div></aside>
    <main class="dash-main"><header class="dash-top"><h1>Admin Dashboard</h1><div>🔔 &nbsp; 👤 <b>{escape(session.get('admin_username','Admin')).title()}</b><br><small>{role_label}</small></div></header><div class="dash-content">
      <section class="welcome"><div><h2>Welcome back, Admin! 👋</h2><span>Here's what's happening in your mess today.</span></div><div class="date-pill">📅 {date_label}</div></section>
      <section class="metric-row">
       <div class="dmetric"><span class="micon">♙</span><div>Total Students<strong>{student_count}</strong><small>Across all hostels</small></div></div>
       <div class="dmetric"><span class="micon">♂</span><div>Boys Hostel<strong>{boys_count}</strong><small>{round((boys_count/student_count*100),1) if student_count else 0}% of total</small></div></div>
       <div class="dmetric"><span class="micon" style="background:#ffe3ef">♀</span><div>Girls Hostel<strong>{girls_count}</strong><small>{round((girls_count/student_count*100),1) if student_count else 0}% of total</small></div></div>
      </section>
      <section class="metric-row small">
       <div class="dmetric"><span class="micon">☕</span><div>Breakfast Today<strong>{breakfast}</strong></div></div><div class="dmetric"><span class="micon">🍲</span><div>Lunch Today<strong>{lunch}</strong></div></div><div class="dmetric"><span class="micon">🍚</span><div>Dinner Today<strong>{dinner}</strong></div></div><div class="dmetric"><span class="micon">♙</span><div>Absent Today<strong>{absent}</strong></div></div><div class="dmetric"><span class="micon">⏭</span><div>Skipped Meals<strong>{skipped}</strong></div></div>
      </section>
      <section class="panel-grid"><div class="dpanel"><h3>Meal Attendance (Last 7 Days)</h3><canvas id="chart7"></canvas></div><div class="dpanel"><h3>Meal Distribution Today</h3><canvas id="donut"></canvas><p class="center"><b>{total_meals}</b> total meals</p></div><div class="dpanel"><h3>Recent Activity</h3>{recent_html}</div></section>
      <section class="dpanel" style="margin-top:18px"><h3>Quick Actions</h3><div class="quick"><a href="/admin/scanner">▣ QR Scan<br><small>Scan student QR</small></a><a href="/admin/students">♙ Students<br><small>Manage records</small></a><a href="/admin/weekly-menu">▦ Weekly Menu<br><small>Manage menu</small></a><a href="/admin/reports">▤ Reports<br><small>Attendance reports</small></a><a href="/admin/complaints">◉ Complaints<br><small>View & resolve</small></a></div></section>
      <p style="text-align:right;color:#60708d">{database_label}</p>
    </div></main><script>
    const c7={chart7};new Chart(document.getElementById('chart7'),{{type:'line',data:{{labels:c7.labels,datasets:[{{label:'Breakfast',data:c7.breakfast,borderColor:'#05aeba',backgroundColor:'transparent',tension:.35}},{{label:'Lunch',data:c7.lunch,borderColor:'#28a745',backgroundColor:'transparent',tension:.35}},{{label:'Dinner',data:c7.dinner,borderColor:'#7c4dff',backgroundColor:'transparent',tension:.35}}]}},options:{{responsive:true,plugins:{{legend:{{position:'top'}}}},scales:{{y:{{beginAtZero:true,ticks:{{precision:0}}}}}}}}}});
    new Chart(document.getElementById('donut'),{{type:'doughnut',data:{{labels:['Breakfast','Lunch','Dinner'],datasets:[{{data:[{breakfast},{lunch},{dinner}],backgroundColor:['#09b7c4','#38aa43','#8453ed']}}]}},options:{{cutout:'65%',plugins:{{legend:{{position:'right'}}}}}}}});
    </script>{PWA_SCRIPT}</body></html>"""

    return html


# =========================================================
# ADMIN - MEAL SETTINGS
# =========================================================

@app.route("/admin/meal-settings", methods=["GET", "POST"])
def admin_meal_settings():
    if not admin_required():
        return redirect(url_for("admin_login"))

    message = ""
    message_class = "message success"
    if request.method == "POST":
        values = {}
        valid = True
        for meal in ["breakfast", "lunch", "dinner"]:
            start = request.form.get(meal + "_start", "").strip()
            end = request.form.get(meal + "_end", "").strip()
            try:
                start_time = datetime.strptime(start, "%H:%M")
                end_time = datetime.strptime(end, "%H:%M")
                if start_time >= end_time:
                    valid = False
            except ValueError:
                valid = False
            values[meal + "_open"] = "1" if request.form.get(meal + "_open") == "1" else "0"
            values[meal + "_start"] = start
            values[meal + "_end"] = end
        if valid:
            save_settings(values)
            message = "Meal settings saved successfully."
        else:
            message = "Please select valid start and end times. End time must be later than start time."
            message_class = "message error"

    settings = {
        meal: {
            "open": get_setting(meal + "_open", "1") == "1",
            "start": get_setting(meal + "_start"),
            "end": get_setting(meal + "_end"),
        }
        for meal in ["breakfast", "lunch", "dinner"]
    }

    cards = ""
    for meal, icon in [("breakfast", "🌅"), ("lunch", "☀️"), ("dinner", "🌙")]:
        item = settings[meal]
        cards += f"""
        <div class="card">
          <h2>{icon} {meal.title()}</h2>
          <label style="display:flex;gap:10px;align-items:center;margin-bottom:18px">
            <input type="checkbox" name="{meal}_open" value="1" {'checked' if item['open'] else ''} style="width:auto;margin:0">
            Open for students
          </label>
          <label>Start Time</label><input type="time" name="{meal}_start" value="{item['start']}" required>
          <label>End Time</label><input type="time" name="{meal}_end" value="{item['end']}" required>
        </div>
        """

    return f"""<!DOCTYPE html><html><head><title>Meal Settings</title>{CSS}</head><body>
    <div class="container">
      <div class="nav"><a class="btn" href="/admin/dashboard">Dashboard</a><a class="btn blue" href="/admin/scanner">Scanner</a></div>
      <h1>⏰ Meal Open/Close & Timings</h1>
      {f'<div class="{message_class}">{escape(message)}</div>' if message else ''}
      <form method="POST"><div class="grid">{cards}</div>
        <div class="card center"><button class="btn green" type="submit">Save Meal Settings</button></div>
      </form>
    </div></body></html>"""


# =========================================================
# ADMIN - ADD STUDENT
# =========================================================


@app.route("/admin/add-student", methods=["GET", "POST"])
def admin_add_student():

    if not admin_required(["MAIN", "BOYS", "GIRLS"]):
        return redirect(url_for("admin_login"))

    message = ""
    message_class = "message"

    if request.method == "POST":

        name = request.form.get("name", "").strip()
        roll = request.form.get("roll_number", "").strip()
        branch = request.form.get("branch", "").strip()
        room = request.form.get("hostel_room", "").strip()
        hostel_name = request.form.get("hostel_name", "").strip()
        if admin_scope_gender(): hostel_name = "Boys Hostel" if admin_scope_gender()=="BOY" else "Girls Hostel"
        gender = "BOY" if hostel_name == "Boys Hostel" else "GIRL" if hostel_name == "Girls Hostel" else ""
        hostel_block = request.form.get("hostel_block", "").strip() if hostel_name=="Boys Hostel" else ""
        pin = request.form.get("pin", "").strip()
        photo = request.files.get("photo")

        if (not name or not roll or branch not in BRANCHES or not room or
                gender not in ["BOY", "GIRL"] or hostel_name not in ["Boys Hostel", "Girls Hostel"] or
                (hostel_name=="Boys Hostel" and hostel_block not in HOSTEL_BLOCKS) or len(pin) != 4 or not pin.isdigit()):

            message = "Please fill every field."
            message_class = "message error"

        elif not photo or not photo.filename:

            message = "Student photo is required."
            message_class = "message error"

        else:

            exists = student_by_roll(roll)

            if exists:

                message = "This Registration Number is already registered."
                message_class = "message error"

            else:

                uid = make_student_uid()

                ext = os.path.splitext(
                    photo.filename
                )[1].lower()

                if ext not in [".jpg", ".jpeg", ".png", ".webp"]:

                    message = "Use JPG, JPEG, PNG or WEBP."
                    message_class = "message error"

                else:

                    filename = uid + ext
                    photo_bytes = photo.read()
                    mime = {".png": "image/png", ".webp": "image/webp"}.get(ext, "image/jpeg")
                    try:
                        add_student_record({
                            "student_uid": uid, "name": name, "roll_number": roll,
                            "branch": branch, "hostel_room": room, "gender": gender,
                            "hostel_name": hostel_name, "hostel_block": hostel_block,
                            "photo_filename": filename, "photo_data": photo_bytes,
                            "photo_mime": mime, "pin_hash": generate_password_hash(pin),
                            "active": 1, "created_at": current_time().strftime("%Y-%m-%d %H:%M:%S")
                        })
                    except (DuplicateKeyError, sqlite3.IntegrityError):
                        message = "This Registration Number is already registered."
                        message_class = "message error"
                        return redirect(url_for("admin_add_student"))

                    message = (
                        "Student registered successfully. "
                        f"Student ID: {uid}"
                    )

                    message_class = "message success"

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Register Student</title>
        {CSS}
    </head>

    <body>
    <div class="container">

        <div class="nav">
            <a class="btn" href="/admin/dashboard">
                Dashboard
            </a>
            <a class="btn blue" href="/admin/students">
                Students
            </a>
        </div>

        <div class="card">

            <h1>👨‍🎓 Register Hostel Student</h1>

            {'<div class="' + message_class + '">' +
             message + '</div>' if message else ''}

            <form
                method="POST"
                enctype="multipart/form-data"
            >

                <label>Student Name</label>

                <input
                    type="text"
                    name="name"
                    required
                >

                <label>Registration Number</label>

                <input
                    type="text"
                    name="roll_number"
                    required
                >

                <label>Branch</label>
                <select name="branch" required>
                    <option value="">Select Branch</option>
                    <option value="AI &amp; ML">AI &amp; ML</option>
                    <option value="Civil (Construction Technology)">Civil (Construction Technology)</option>
                    <option value="Electronics (Robotics)">Electronics (Robotics)</option>
                    <option value="Mechanical (CAD/CAM)">Mechanical (CAD/CAM)</option>
                </select>

                <label>Hostel Room</label>

                <input
                    type="text"
                    name="hostel_room"
                    required
                >

                <label>Hostel</label>
                <select name="hostel_name" id="hostel" onchange="toggleBlock()" required>
                    <option value="">Select Hostel</option>
                    <option value="Boys Hostel">Boys Hostel</option>
                    <option value="Girls Hostel">Girls Hostel</option>
                </select>

                <div id="blockWrap"><label>Hostel Block (Boys Hostel only)</label>
                <select name="hostel_block" id="block">
                    <option value="">Select Hostel Block</option>
                    <option value="BH-1">BH-1</option>
                    <option value="BH-2">BH-2</option>
                </select></div>

                <label>4-digit Student PIN</label>
                <input type="password" name="pin" inputmode="numeric" minlength="4" maxlength="4" pattern="[0-9]{{4}}" required>

                <label>Student Photo</label>

                <input
                    type="file"
                    name="photo"
                    accept="image/*"
                    required
                >

                <button class="btn green" type="submit">
                    Register Student
                </button>

            </form>

        </div>

    </div>
    <script>function toggleBlock(){{const boys=document.getElementById('hostel').value==='Boys Hostel';document.getElementById('blockWrap').style.display=boys?'block':'none';document.getElementById('block').required=boys;if(!boys)document.getElementById('block').value='';}}toggleBlock();</script></body>
    </html>
    """

    return html


# =========================================================
# ADMIN - STUDENT LIST
# =========================================================

@app.route("/admin/students")
def admin_students():

    if not admin_required(["MAIN", "BOYS", "GIRLS"]):
        return redirect(url_for("admin_login"))

    search = request.args.get("q", "").strip()
    gender_filter = request.args.get("gender", "").strip().upper()
    if admin_scope_gender(): gender_filter = admin_scope_gender()
    hostel_filter = request.args.get("hostel", "").strip()
    branch_filter = request.args.get("branch", "").strip()
    block_filter = request.args.get("block", "").strip()
    students = list_student_records(
        search=search,
        gender=gender_filter if gender_filter in ["BOY", "GIRL"] else "",
        hostel=hostel_filter if hostel_filter in ["Boys Hostel", "Girls Hostel"] else "",
        branch=branch_filter if branch_filter in BRANCHES else "",
        block=block_filter if block_filter in HOSTEL_BLOCKS else "",
    )

    rows = ""

    for student in students:

        photo = "No Photo"

        if student.get("photo_filename") or student.get("photo_data"):

            photo = f"""
                <img
                    class="photo"
                    src="/student-photo/{student['id']}"
                >
            """

        status = (
            '<span class="badge active-badge">ACTIVE</span>'
            if student["active"]
            else
            '<span class="badge expired-badge">INACTIVE</span>'
        )

        rows += f"""
        <tr>

            <td>{photo}</td>
            <td>{escape(student["name"])}</td>
            <td>{escape(student["roll_number"])}</td>
            <td>{escape(student["gender"] or "NOT SET")}</td>
            <td>{escape(student["branch"])}</td>
            <td>{escape(student["hostel_name"] or "NOT SET")}<br><small>{escape(student["hostel_block"] or "")}</small></td>
            <td>{escape(student["hostel_room"])}</td>
            <td>{student["student_uid"]}</td>
            <td>{status}</td>
            <td><a class="btn blue" href="/admin/student/{student['id']}/edit">Edit</a>
                <form method="POST" action="/admin/student/{student['id']}/toggle" style="display:inline">
                  <button class="btn {'red' if student['active'] else 'green'}" type="submit">{'Inactive' if student['active'] else 'Active'}</button>
                </form></td>

        </tr>
        """

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Students</title>
        {CSS}
    </head>

    <body>

    <div class="container">

        <div class="nav">

            <a class="btn" href="/admin/dashboard">
                Dashboard
            </a>

            <a class="btn green" href="/admin/add-student">
                Register Student
            </a>

        </div>

        <div class="card">

            <h1>👨‍🎓 Hostel Students</h1>

            <form method="GET" class="grid" style="align-items:end">
              <div><label>Search</label><input name="q" value="{escape(search)}" placeholder="Name / Registration No. / Branch"></div>
              <div><label>Gender</label><select name="gender"><option value="">All</option><option value="BOY" {'selected' if gender_filter == 'BOY' else ''}>Boys</option><option value="GIRL" {'selected' if gender_filter == 'GIRL' else ''}>Girls</option></select></div>
              <div><label>Hostel</label><select name="hostel"><option value="">All</option><option value="Boys Hostel" {'selected' if hostel_filter == 'Boys Hostel' else ''}>Boys Hostel</option><option value="Girls Hostel" {'selected' if hostel_filter == 'Girls Hostel' else ''}>Girls Hostel</option></select></div>
              <div><label>Branch</label><select name="branch"><option value="">All Branches</option>{''.join(f'<option value="{escape(branch)}" {"selected" if branch_filter == branch else ""}>{escape(branch)}</option>' for branch in BRANCHES)}</select></div>
              <div><label>Block</label><select name="block"><option value="">All Blocks</option>{''.join(f'<option value="{block}" {"selected" if block_filter == block else ""}>{block}</option>' for block in HOSTEL_BLOCKS)}</select></div>
              <div><button class="btn blue" type="submit">Apply Filter</button> <a class="btn gray" href="/admin/students">Clear</a></div>
            </form>

            <div style="overflow-x:auto;">

                <table>

                    <tr>
                        <th>Photo</th>
                        <th>Name</th>
                        <th>Registration No.</th>
                        <th>Gender</th>
                        <th>Branch</th>
                        <th>Hostel/Block</th>
                        <th>Room</th>
                        <th>Student ID</th>
                        <th>Status</th>
                        <th>Action</th>
                    </tr>

                    {rows}

                </table>

            </div>

        </div>

    </div>

    </body>
    </html>
    """

    return html


@app.route("/admin/student/<int:student_id>/edit", methods=["GET", "POST"])
def admin_edit_student(student_id):
    if not admin_required(["MAIN", "BOYS", "GIRLS"]):
        return redirect(url_for("admin_login"))

    student = student_by_id(student_id)
    if not student:
        return "Student not found", 404
    if admin_scope_gender() and student.get("gender") != admin_scope_gender():
        return "Not allowed for this hostel admin.", 403

    message = ""
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        registration = request.form.get("registration_number", "").strip()
        branch = request.form.get("branch", "").strip()
        room = request.form.get("hostel_room", "").strip()
        hostel_name = request.form.get("hostel_name", "").strip()
        if admin_scope_gender(): hostel_name = "Boys Hostel" if admin_scope_gender()=="BOY" else "Girls Hostel"
        gender = "BOY" if hostel_name=="Boys Hostel" else "GIRL" if hostel_name=="Girls Hostel" else ""
        hostel_block = request.form.get("hostel_block", "").strip() if hostel_name=="Boys Hostel" else ""
        pin = request.form.get("pin", "").strip()
        photo = request.files.get("photo")
        if (not name or not registration or branch not in BRANCHES or not room or
                gender not in ["BOY", "GIRL"] or hostel_name not in ["Boys Hostel", "Girls Hostel"] or
                (hostel_name=="Boys Hostel" and hostel_block not in HOSTEL_BLOCKS)):
            message = "Please fill all required fields."
        elif pin and (len(pin) != 4 or not pin.isdigit()):
            message = "New PIN must contain exactly 4 numbers."
        else:
            duplicate = student_by_roll(registration)
            if duplicate and duplicate["id"] != student_id:
                message = "Registration Number already exists."
            else:
                changes = {
                    "name": name,
                    "roll_number": registration,
                    "branch": branch,
                    "hostel_room": room,
                    "gender": gender,
                    "hostel_name": hostel_name,
                    "hostel_block": hostel_block,
                }
                if photo and photo.filename:
                    ext = os.path.splitext(photo.filename)[1].lower()
                    if ext not in [".jpg", ".jpeg", ".png", ".webp"]:
                        message = "Use JPG, JPEG, PNG or WEBP photo."
                    else:
                        changes.update({
                            "photo_filename": student["student_uid"] + ext,
                            "photo_data": photo.read(),
                            "photo_mime": {".png": "image/png", ".webp": "image/webp"}.get(ext, "image/jpeg"),
                        })
                if pin:
                    changes["pin_hash"] = generate_password_hash(pin)
                if not message:
                    update_student_record(student_id, changes)
                    return redirect(url_for("admin_students"))

    student = student_by_id(student_id)
    html = f"""<!DOCTYPE html><html><head><title>Edit Student</title>{CSS}</head><body><div class="container">
    <div class="nav"><a class="btn" href="/admin/students">Back to Students</a></div>
    <div class="card"><h1>✏️ Edit Student</h1>{f'<div class="message error">{escape(message)}</div>' if message else ''}
    <form method="POST" enctype="multipart/form-data">
      <label>Name</label><input name="name" value="{escape(student['name'])}" required>
      <label>Registration Number</label><input name="registration_number" value="{escape(student['roll_number'])}" required>
      <label>Branch</label><select name="branch" required>
        {''.join(f'<option value="{escape(branch)}" {"selected" if student["branch"] == branch else ""}>{escape(branch)}</option>' for branch in BRANCHES)}
      </select>
      <label>Hostel</label><select name="hostel_name" id="hostel" onchange="toggleBlock()" required><option value="Boys Hostel" {'selected' if student['hostel_name']=='Boys Hostel' else ''}>Boys Hostel</option><option value="Girls Hostel" {'selected' if student['hostel_name']=='Girls Hostel' else ''}>Girls Hostel</option></select>
      <div id="blockWrap"><label>Hostel Block (Boys Hostel only)</label><select name="hostel_block" id="block">
        {''.join(f'<option value="{block}" {"selected" if student["hostel_block"] == block else ""}>{block}</option>' for block in HOSTEL_BLOCKS)}
      </select></div>
      <label>Room Number</label><input name="hostel_room" value="{escape(student['hostel_room'])}" required>
      <label>Reset 4-digit PIN (optional)</label><input type="password" name="pin" inputmode="numeric" minlength="4" maxlength="4" pattern="[0-9]{{4}}" placeholder="Leave blank to keep current PIN">
      <label>Change Photo (optional)</label><input type="file" name="photo" accept="image/*">
      <button class="btn green" type="submit">Save Changes</button>
    </form></div></div><script>function toggleBlock(){{const boys=document.getElementById('hostel').value==='Boys Hostel';document.getElementById('blockWrap').style.display=boys?'block':'none';document.getElementById('block').required=boys;if(!boys)document.getElementById('block').value='';}}toggleBlock();</script></body></html>"""
    return html


@app.route("/admin/student/<int:student_id>/toggle", methods=["POST"])
def admin_toggle_student(student_id):
    if not admin_required(["MAIN", "BOYS", "GIRLS"]):
        return redirect(url_for("admin_login"))
    student = student_by_id(student_id)
    if admin_scope_gender() and student and student.get("gender") != admin_scope_gender():
        return "Not allowed for this hostel admin.", 403
    toggle_student_record(student_id)
    return redirect(url_for("admin_students"))


# =========================================================
# STUDENT PHOTO
# =========================================================

@app.route("/student-photo/<int:student_id>")
def student_photo(student_id):
    student = student_by_id(student_id)
    if not student:
        return "Photo not found", 404
    if student.get("photo_data"):
        return Response(bytes(student["photo_data"]), mimetype=student.get("photo_mime") or "image/jpeg")
    filename = student.get("photo_filename")
    if filename and os.path.exists(os.path.join(PHOTO_DIR, filename)):
        return send_from_directory(PHOTO_DIR, filename)
    return "Photo not found", 404


# =========================================================
# STUDENT PANEL
# =========================================================

@app.route("/student")
def student_home():

    # Student session is not required permanently.
    # Registration number verification is done before coupon generation.

    html = f"""<!doctype html><html><head><title>Student Login | SmartMess</title>{CSS}{DASHBOARD_CSS}{PWA_HEAD}</head><body class="login-page">
    <main class="login-shell"><section class="login-hero"><div class="login-logo"><img src="/static/icon-192.png" alt="College logo"><div><strong>Government Polytechnic, Barh</strong><span>SmartMess Management System</span></div></div><div class="login-hero-content"><h1>SmartMess</h1><p>Student Meal Services</p><div class="login-features"><div class="login-feature"><i><svg viewBox="0 0 24 24"><rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/><rect x="3" y="14" width="7" height="7"/><path d="M14 14h3v3h4v4h-7z"/></svg></i> Generate QR Coupon</div><div class="login-feature"><i><svg viewBox="0 0 24 24"><path d="M4 11h16M5 11a7 7 0 0 1 14 0M3 15h18M8 19h8"/><path d="M12 4V2"/></svg></i> Today’s Menu</div><div class="login-feature"><i><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="9"/><path d="M12 7v5l3 2M4 4l2 2M20 4l-2 2"/></svg></i> Meal History &amp; Skip Meal</div></div></div><div class="login-building">GOVERNMENT POLYTECHNIC · BARH</div></section>
    <section class="login-pane"><div class="login-box"><h2>Welcome to SmartMess</h2><p class="sub">Student meal access</p><nav class="login-tabs"><a class="on" href="/student"><span class="tab-icon"><svg viewBox="0 0 24 24"><path d="m3 10 9-5 9 5-9 5zM7 12v5c3 2 7 2 10 0v-5M21 10v6"/></svg></span>Student</a><a href="/admin"><span class="tab-icon"><svg viewBox="0 0 24 24"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/><circle cx="12" cy="9" r="2.2"/><path d="M8.5 16c.7-2 2-3 3.5-3s2.8 1 3.5 3"/></svg></span>Admin</a></nav><form method="post" action="/student/verify"><label>Registration Number</label><input name="roll_number" placeholder="Enter your registration number" autocomplete="username" required><label>4-digit PIN</label><div class="field-wrap"><input id="loginSecret" type="password" name="pin" inputmode="numeric" minlength="4" maxlength="4" pattern="[0-9]{{4}}" placeholder="Enter 4-digit PIN" autocomplete="current-password" required><button type="button" class="eye-btn" onclick="toggleSecret()" aria-label="Show PIN"><svg viewBox="0 0 24 24" width="21" height="21" fill="none" stroke="currentColor" stroke-width="2"><path d="M2 12s3.5-7 10-7 10 7 10 7-3.5 7-10 7S2 12 2 12z"/><circle cx="12" cy="12" r="3"/></svg></button></div><button class="login-primary" type="submit">Login to Student Dashboard</button></form><div class="login-help"><a href="/student/forgot-pin">Forgot PIN?</a></div><div class="security-note"><span class="security-icon"><svg viewBox="0 0 24 24"><path d="M12 22s8-4 8-10V5l-8-3-8 3v7c0 6 8 10 8 10z"/><path d="m9 12 2 2 4-4"/></svg></span><span>Only registered and active hostel students can access meal coupons.</span></div></div></section></main>
    <script>function toggleSecret(){{const x=document.getElementById('loginSecret');x.type=x.type==='password'?'text':'password'}};</script>{PWA_SCRIPT}</body></html>"""

    return html


@app.route("/student/verify", methods=["POST"])
def student_verify():

    roll = request.form.get(
        "roll_number",
        ""
    ).strip()
    pin = request.form.get("pin", "").strip()
    key="student:"+roll
    if auth_locked(key):
        return f'''<!doctype html><html><head>{CSS}</head><body><div class="container"><div class="card center"><h1>🔒 Login Locked</h1><div class="message error">Too many wrong attempts. Try again after 15 minutes.</div><a class="btn" href="/student">Back</a></div></div></body></html>'''
    student = student_by_roll(roll, active_only=True)

    if not student or not student.get("pin_hash") or not check_password_hash(student["pin_hash"], pin):

        auth_fail(key)
        return f"""
        <!DOCTYPE html>
        <html>
        <head>{CSS}</head>
        <body>

        <div class="container">

            <div class="card center">

                <h1>❌ Login Failed</h1>

                <div class="message error">
                    Registration Number or PIN is incorrect, or this student is inactive.
                </div>

                <a class="btn" href="/student">
                    Try Again
                </a>

                </div>

        </div>

        </body>
        </html>
        """

    auth_clear(key); session["student_uid"] = student["student_uid"]; session.permanent=True
    update_student_record(student["id"],{"last_login":current_time().strftime("%Y-%m-%d %H:%M:%S")})
    if student.get("force_pin_change"):
        return redirect(url_for("student_change_pin"))

    return redirect(
        url_for(
            "student_meals"
        )
    )


@app.route("/student/forgot-pin",methods=["GET","POST"])
def student_forgot_pin():
    message=""
    if request.method=="POST":
        roll=request.form.get("roll_number","").strip(); room=request.form.get("hostel_room","").strip(); reason=request.form.get("reason","").strip()
        student=student_by_roll(roll,active_only=True)
        if student and hmac.compare_digest(str(student.get("hostel_room","")),room): add_pin_request(student["student_uid"],room,reason)
        message="Request submitted. Contact your hostel admin for the temporary PIN."
    return f'''<!doctype html><html><head><title>Forgot PIN</title>{CSS}</head><body><div class="container"><div class="card" style="max-width:560px;margin:60px auto"><h1>🔑 Forgot PIN Request</h1>{f'<div class="message success">{message}</div>' if message else ''}<form method="post"><label>Registration Number</label><input name="roll_number" required><label>Hostel Room</label><input name="hostel_room" required><label>Reason</label><input name="reason" maxlength="200"><button class="btn blue">Send Request</button> <a class="btn gray" href="/student">Back</a></form></div></div></body></html>'''


@app.route("/student/change-pin",methods=["GET","POST"])
def student_change_pin():
    uid=session.get("student_uid")
    if not uid:return redirect(url_for("student_home"))
    student=student_by_uid(uid,active_only=True); message=""; ok=False
    if request.method=="POST":
        old=request.form.get("old_pin",""); new=request.form.get("new_pin",""); confirm=request.form.get("confirm_pin","")
        if not check_password_hash(student["pin_hash"],old): message="Current PIN is incorrect."
        elif not re.fullmatch(r"\d{4}",new) or new in ["0000","1111","1234","4321"]: message="Choose a stronger 4-digit PIN."
        elif new!=confirm: message="New PIN confirmation does not match."
        else:
            update_student_record(student["id"],{"pin_hash":generate_password_hash(new),"pin_changed_at":current_time().strftime("%Y-%m-%d %H:%M:%S"),"force_pin_change":0}); ok=True; message="PIN changed successfully."
    return f'''<!doctype html><html><head><title>Change PIN</title>{CSS}</head><body><div class="container"><div class="card" style="max-width:560px;margin:50px auto"><h1>🔐 Change PIN</h1>{f'<div class="message {"success" if ok else "error"}">{escape(message)}</div>' if message else ''}<form method="post"><label>Current PIN</label><input type="password" name="old_pin" pattern="[0-9]{{4}}" required><label>New 4-digit PIN</label><input type="password" name="new_pin" pattern="[0-9]{{4}}" required><label>Confirm New PIN</label><input type="password" name="confirm_pin" pattern="[0-9]{{4}}" required><button class="btn green">Change PIN</button> <a class="btn gray" href="/student/meals">Back</a></form></div></div></body></html>'''


# =========================================================
# STUDENT MEAL PAGE
# =========================================================

@app.route("/student/meals")
def student_meals():

    student_uid = session.get("student_uid")

    if not student_uid:
        return redirect(url_for("student_home"))

    student = student_by_uid(student_uid, active_only=True)

    if not student:

        session.pop("student_uid", None)

        return redirect(
            url_for("student_home")
        )

    photo = f"/student-photo/{student['id']}" if student.get("photo_filename") or student.get("photo_data") else ""

    meal_status_cards = ""
    for meal, icon in [("BREAKFAST", "🌅"), ("LUNCH", "☀️"), ("DINNER", "🌙")]:
        available, status_message = meal_is_available(meal)
        if available:
            status_html = '<span class="badge used-badge">OPEN NOW</span>'
        else:
            status_html = '<span class="badge expired-badge">CLOSED</span>'
        meal_status_cards += (
            f'<div class="stat"><div>{icon} {meal.title()}</div>'
            f'<p>{status_html}</p><small>{escape(status_message) if status_message else "Coupon available"}</small></div>'
        )

    today = current_time().strftime("%Y-%m-%d")
    month_start = current_time().strftime("%Y-%m-01")
    history = list_used_records(month_start, today, registration=student["roll_number"])
    history_rows = "".join(f"<tr><td>{(r.get('used_at') or '')[:10]}</td><td>{r.get('meal','-')}</td><td>{(r.get('used_at') or '')[11:19]}</td></tr>" for r in history[:31])
    today_menu=weekly_menu().get(current_time().strftime("%A"),{})
    menus = {meal: today_menu.get(meal) or get_setting(meal.lower() + "_menu", "Menu not updated") or "Menu not updated" for meal in ["BREAKFAST", "LUNCH", "DINNER"]}
    notice = get_setting("student_notice", "")
    tomorrow = (current_time().date() + timedelta(days=1)).isoformat()
    skip_rows=""
    active_skips=list_skips(student_uid,today,tomorrow)
    for sk in active_skips:
        can_cancel=sk['skip_date']>today or current_time().strftime('%H:%M')<get_setting(sk['meal'].lower()+'_start','00:00')
        action=f'''<form method="post" action="/student/cancel-skip"><input type="hidden" name="meal" value="{sk['meal']}"><input type="hidden" name="skip_date" value="{sk['skip_date']}"><button class="btn gray">Cancel Skip</button></form>''' if can_cancel else 'Meal started'
        skip_rows+=f"<tr><td>{sk['skip_date']}</td><td>{sk['meal']}</td><td>{action}</td></tr>"
    upcoming=active_skips[0] if active_skips else None
    upcoming_html=(f'''<h3>{escape(upcoming['meal'].title())} · {escape(upcoming['skip_date'])}</h3><span class="badge" style="background:#fff0c9;color:#b76a00">SKIPPED</span><form method="post" action="/student/cancel-skip" style="margin-top:14px"><input type="hidden" name="meal" value="{upcoming['meal']}"><input type="hidden" name="skip_date" value="{upcoming['skip_date']}"><button class="btn red" style="width:100%">Cancel Skip</button></form>''' if upcoming else '<p>No upcoming skipped meal.</p>')
    recent_meals="".join(f'''<div class="activity"><span class="micon" style="width:38px;height:38px;font-size:18px">🍴</span><div style="flex:1"><b>{escape(r.get('meal','').title())}</b><br><small>{escape((r.get('used_at') or '')[:16])}</small></div><span class="status-open">PRESENT</span></div>''' for r in history[:5]) or '<p>No meal history yet.</p>'
    month_days=[]; present_series=[]; skipped_series=[]
    for day in [1,5,10,15,20,25,current_time().day]:
        day=min(day,current_time().day)
        if day in month_days: continue
        month_days.append(day)
        cutoff=f"{current_time().strftime('%Y-%m')}-{day:02d}"
        present_series.append(sum(1 for r in history if (r.get('used_at') or '')[:10]<=cutoff))
        skipped_series.append(sum(1 for s in list_skips(student_uid,month_start,today) if s.get('skip_date','')<=cutoff))
    profile_photo=f'<img src="{photo}" alt="Student photo">' if photo else '<div class="micon" style="width:122px;height:122px;font-size:55px">👤</div>'
    date_label=current_time().strftime("%d %b %Y, %A")
    weekly=weekly_menu()
    weekly_rows="".join(f'<tr><td><b>{day}</b></td><td>{escape(items.get("BREAKFAST","") or "-")}</td><td>{escape(items.get("LUNCH","") or "-")}</td><td>{escape(items.get("DINNER","") or "-")}</td></tr>' for day,items in weekly.items())

    html=f"""<!doctype html><html><head><title>Student Dashboard</title>{CSS}{DASHBOARD_CSS}{PWA_HEAD}<script src="https://cdn.jsdelivr.net/npm/chart.js"></script></head><body class="dash student-layout">
    <aside class="side"><div class="brand" style="display:block;text-align:center"><img src="/static/icon-192.png" style="width:78px;height:78px"><p><b>{COLLEGE_NAME}</b></p></div><div class="side-title">🍽️ SmartMess</div><nav><a class="on" href="/student/meals">⌂ Dashboard</a><a href="#generate">▣ Generate QR Coupon</a><a href="#skip">⊘ Skip Meal</a><a href="#history">◷ Meal History</a><a href="#menu">▦ Weekly Menu</a><a href="/student/complaints">▣ Complaints & Feedback</a><a href="/student/change-pin">♢ Change PIN</a><a href="/student/logout">↪ Logout</a></nav><div class="side-foot">◷ Last Login<br><b>{escape(student.get('last_login') or 'First login')}</b></div></aside>
    <main class="dash-main"><header class="dash-top"><h1>☰ &nbsp; Student Dashboard</h1><div><button class="install-pwa" onclick="installSmartMess()">⇩ Install App</button> &nbsp; 🔔 Notice &nbsp; 👤 <b>{escape(student['name'].split()[0])}</b></div></header><div class="dash-content">
      <h2>Welcome, <span style="color:#075ee8">{escape(student['name'].split()[0])}</span> 👋</h2>
      <section class="metric-row"><div class="dmetric profile-box">{profile_photo}<div><h2>{escape(student['name'])} <span class="status-open" style="font-size:12px">● Active</span></h2><p>▣ Registration No. {escape(student['roll_number'])}</p><p>⌂ {escape(student.get('hostel_name') or '')} {escape(student.get('hostel_block') or '')}</p><p>▯ Room {escape(student['hostel_room'])}</p></div></div><div id="generate" class="coupon-call"><div class="fakeqr">▦</div><div><h2>Generate QR Coupon</h2><p style="color:#e6f1ff">Scan at the mess to get your meal</p><a class="btn" href="#menu">Choose Meal ↓</a></div></div></section>
      <section id="menu" style="margin-top:22px"><div class="welcome"><h2>Today's Menu</h2><a href="#weekly">View Weekly Menu ›</a></div><div class="menu-grid">
       <div class="menu-card"><h3>🌅 Breakfast</h3><p>{escape(menus['BREAKFAST'])}</p><p>{get_setting('breakfast_start')} – {get_setting('breakfast_end')}</p><form method="post" action="/student/generate"><input type="hidden" name="meal" value="BREAKFAST"><button class="btn green">Generate Breakfast QR</button></form></div>
       <div class="menu-card"><h3>☀️ Lunch</h3><p>{escape(menus['LUNCH'])}</p><p>{get_setting('lunch_start')} – {get_setting('lunch_end')}</p><form method="post" action="/student/generate"><input type="hidden" name="meal" value="LUNCH"><button class="btn blue">Generate Lunch QR</button></form></div>
       <div class="menu-card"><h3>🌙 Dinner</h3><p>{escape(menus['DINNER'])}</p><p>{get_setting('dinner_start')} – {get_setting('dinner_end')}</p><form method="post" action="/student/generate"><input type="hidden" name="meal" value="DINNER"><button class="btn" style="background:#7948df">Generate Dinner QR</button></form></div>
      </div></section>
      <section class="student-lower"><div class="dpanel"><h3>Monthly Attendance</h3><canvas id="monthChart"></canvas></div><div id="history" class="dpanel"><h3>Recent Meal History</h3>{recent_meals}</div><div><div class="dpanel"><h3>Upcoming Skipped Meal</h3>{upcoming_html}</div><div id="skip" class="dpanel" style="margin-top:15px"><h3>Skip a Meal</h3><form method="post" action="/student/skip-meal"><select name="meal"><option>BREAKFAST</option><option>LUNCH</option><option>DINNER</option></select><select name="skip_date"><option value="{today}">Today</option><option value="{tomorrow}">Tomorrow</option></select><button class="btn red" style="width:100%">Skip Meal</button></form></div></div></section>
      <section class="dpanel" style="margin-top:18px"><h3>Quick Actions</h3><div class="quick"><a href="#menu">▣ Generate QR</a><a href="#skip">⊘ Skip Meal</a><a href="#history">◷ Meal History</a><a href="/student/complaints">▣ Feedback</a><a href="/student/change-pin">♢ Change PIN</a></div></section>
      <section id="weekly" class="dpanel" style="margin-top:18px"><h3>Weekly Menu</h3><div style="overflow:auto"><table><tr><th>Day</th><th>Breakfast</th><th>Lunch</th><th>Dinner</th></tr>{weekly_rows}</table></div></section>
      {f'<div class="notice-bar">📣 <b>Notice:</b> {escape(notice)}</div>' if notice else ''}
    </div></main><script>new Chart(document.getElementById('monthChart'),{{type:'line',data:{{labels:{json.dumps([str(d) for d in month_days])},datasets:[{{label:'Present',data:{json.dumps(present_series)},borderColor:'#16a85a',tension:.3}},{{label:'Skipped',data:{json.dumps(skipped_series)},borderColor:'#fa4965',tension:.3}}]}},options:{{responsive:true,scales:{{y:{{beginAtZero:true,ticks:{{precision:0}}}}}}}}}});</script>{PWA_SCRIPT}</body></html>"""

    return html


# =========================================================
# STUDENT GENERATE COUPON
# =========================================================

@app.route("/student/generate", methods=["POST"])
def student_generate():

    student_uid = session.get("student_uid")

    if not student_uid:

        return redirect(
            url_for("student_home")
        )

    meal = request.form.get(
        "meal",
        ""
    ).upper()

    if meal not in [
        "BREAKFAST",
        "LUNCH",
        "DINNER"
    ]:

        return "Invalid meal.", 400

    student = student_by_uid(student_uid, active_only=True)

    if not student:
        return redirect(
            url_for("student_home")
        )

    available, unavailable_message = meal_is_available(meal)
    if not available:
        return f"""
        <!DOCTYPE html><html><head><title>Meal Closed</title>{CSS}</head><body>
        <div class="container"><div class="card center">
          <h1>⏰ {meal.title()} Closed</h1>
          <div class="message error">{escape(unavailable_message)}</div>
          <a class="btn" href="/student/meals">Back to Meals</a>
        </div></div></body></html>
        """

    # -----------------------------------------------------
    # Check today's same meal
    # -----------------------------------------------------

    today = current_time().strftime("%Y-%m-%d")

    if any(item["meal"] == meal for item in list_skips(student_uid, today, today)):
        return f"""<!DOCTYPE html><html><head><title>Meal Skipped</title>{CSS}</head><body><div class="container"><div class="card center">
        <h1>⏭️ Meal Skipped</h1><div class="message error">You marked today's {meal.title()} as skipped.</div><a class="btn" href="/student/meals">Back</a>
        </div></div></body></html>"""

    existing = latest_coupon(student_uid, meal, today)

    if existing:

        if existing["status"] == "USED":

            return f"""
            <!DOCTYPE html>
            <html>
            <head>{CSS}</head>
            <body>

            <div class="container">

                <div class="card center">

                    <h1>❌ Already Used</h1>

                    <div class="message error">

                        Your {meal.title()} coupon
                        has already been used today.

                    </div>

                    <a
                        class="btn"
                        href="/student/meals"
                    >
                        Back
                    </a>

                </div>

            </div>

            </body>
            </html>
            """

        if existing["status"] == "ACTIVE":

            expiry = datetime.strptime(
                existing["expires_at"],
                "%Y-%m-%d %H:%M:%S"
            )

            if current_time() < expiry:

                # Show existing active coupon.
                return render_coupon(
                    student,
                    existing
                )

            else:

                update_coupon(existing["id"], {"status": "EXPIRED"})

    # -----------------------------------------------------
    # Create new coupon
    # -----------------------------------------------------

    generated = current_time()

    expires = generated + timedelta(minutes=5)

    token = make_coupon_token()

    coupon = add_coupon_record({
        "token": token,
        "student_uid": student_uid,
        "meal": meal,
        "generated_at": generated.strftime("%Y-%m-%d %H:%M:%S"),
        "expires_at": expires.strftime("%Y-%m-%d %H:%M:%S"),
        "used_at": None,
        "status": "ACTIVE",
    })

    return render_coupon(
        student,
        coupon
    )


@app.route("/student/skip-meal", methods=["POST"])
def student_skip_meal():
    uid = session.get("student_uid")
    if not uid: return redirect(url_for("student_home"))
    meal, skip_date = request.form.get("meal", "").upper(), request.form.get("skip_date", "")
    allowed_dates = {current_time().date().isoformat(), (current_time().date() + timedelta(days=1)).isoformat()}
    if meal not in ["BREAKFAST", "LUNCH", "DINNER"] or skip_date not in allowed_dates: return "Invalid meal or date.", 400
    success = add_skip_record(uid, meal, skip_date)
    text = "Skip meal saved. Thank you for helping reduce food waste." if success else "This meal is already marked as skipped."
    return f"""<!DOCTYPE html><html><head><title>Skip Meal</title>{CSS}</head><body><div class="container"><div class="card center"><h1>⏭️ Skip Meal</h1><div class="message {'success' if success else 'error'}">{text}</div><a class="btn" href="/student/meals">Back to Meals</a></div></div></body></html>"""


@app.route("/student/cancel-skip",methods=["POST"])
def student_cancel_skip():
    uid=session.get("student_uid"); meal=request.form.get("meal","").upper(); day=request.form.get("skip_date","")
    if not uid:return redirect(url_for("student_home"))
    today=current_time().date().isoformat()
    if meal not in ["BREAKFAST","LUNCH","DINNER"] or day<today:return "Cancellation not allowed.",400
    if day==today and current_time().strftime("%H:%M")>=get_setting(meal.lower()+"_start","00:00"): return "Meal start time has passed; skip cannot be cancelled.",400
    delete_skip(uid,meal,day)
    return redirect(url_for("student_meals"))


@app.route("/student/complaints",methods=["GET","POST"])
def student_complaints():
    uid=session.get("student_uid")
    if not uid:return redirect(url_for("student_home"))
    if request.method=="POST":
        category=request.form.get("category",""); meal=request.form.get("meal",""); message=request.form.get("message","").strip(); rating=int(request.form.get("rating","0") or 0)
        if category in ["Food Quality","Hygiene","Service","Other"] and meal in ["BREAKFAST","LUNCH","DINNER","GENERAL"] and 1<=rating<=5 and 5<=len(message)<=500: add_complaint(uid,meal,category,rating,message)
    rows="".join(f'<tr><td>{escape(c["created_at"][:10])}</td><td>{escape(c["category"])}</td><td>{"★"*int(c.get("rating",0))}</td><td>{escape(c["message"])}</td><td>{escape(c.get("status","OPEN"))}<br><small>{escape(c.get("admin_reply","") or "")}</small></td></tr>' for c in list_complaints(uid))
    return f'''<!doctype html><html><head><title>Feedback</title>{CSS}</head><body><div class="container"><div class="nav"><a class="btn" href="/student/meals">Dashboard</a></div><div class="card"><h1>💬 Complaint & Feedback</h1><form method="post"><div class="grid"><div><label>Category</label><select name="category"><option>Food Quality</option><option>Hygiene</option><option>Service</option><option>Other</option></select></div><div><label>Meal</label><select name="meal"><option>GENERAL</option><option>BREAKFAST</option><option>LUNCH</option><option>DINNER</option></select></div><div><label>Rating</label><select name="rating"><option value="5">5 - Excellent</option><option value="4">4 - Good</option><option value="3">3 - Average</option><option value="2">2 - Poor</option><option value="1">1 - Bad</option></select></div></div><label>Message</label><input name="message" minlength="5" maxlength="500" required><button class="btn blue">Submit</button></form></div><div class="card"><h2>My Requests</h2><div style="overflow:auto"><table><tr><th>Date</th><th>Category</th><th>Rating</th><th>Message</th><th>Status / Reply</th></tr>{rows or '<tr><td colspan="5">No feedback yet.</td></tr>'}</table></div></div></div></body></html>'''


# =========================================================
# COUPON HTML
# =========================================================

def render_coupon(student, coupon):
    photo = f"/student-photo/{student['id']}" if student.get("photo_filename") or student.get("photo_data") else ""

    qr = qr_data_uri(
        coupon["token"]
    )

    expires = datetime.strptime(
        coupon["expires_at"],
        "%Y-%m-%d %H:%M:%S"
    )

    expires_iso = expires.isoformat()

    meal_name = coupon["meal"].title()

    html = f"""
    <!DOCTYPE html>
    <html>

    <head>

        <title>Mess Coupon</title>

        {CSS}

    </head>

    <body>

    <div class="container">

        <div class="card center">

            <h1>🎫 Meal Coupon</h1>

            {(
                '<img class="big-photo" src="' +
                photo +
                '">'
            ) if photo else ''}

            <h2>
                {student["name"]}
            </h2>

            <p>
                Registration Number:
                <b>{student["roll_number"]}</b>
            </p>

            <p>
                Branch:
                <b>{student["branch"]}</b>
            </p>

            <p>
                Hostel Room:
                <b>{student["hostel_room"]}</b>
            </p>

            <p>Gender: <b>{student["gender"] or "NOT SET"}</b></p>
            <p>Hostel: <b>{student["hostel_name"] or "NOT SET"} {student["hostel_block"] or ""}</b></p>

            <h2>
                🍽️ {meal_name}
            </h2>

            <img
                class="qr"
                src="{qr}"
                alt="Secure Coupon QR"
            >

            <p>
                Show this QR at the mess scanner.
            </p>

            <div
                id="timer"
                class="countdown"
            >
                05:00
            </div>

            <p>
                Valid for 5 minutes only.
            </p>

            <p>
                After successful scan this coupon
                becomes USED.
            </p>

        </div>

    </div>


    <script>

    const expiry =
        new Date(
            "{expires_iso}"
        ).getTime();

    const timerElement =
        document.getElementById("timer");

    const interval =
        setInterval(function() {{

            const now =
                new Date().getTime();

            const distance =
                expiry - now;

            if (distance <= 0) {{

                clearInterval(interval);

                timerElement.innerText =
                    "EXPIRED";

                timerElement.style.color =
                    "red";

                return;

            }}

            const minutes =
                Math.floor(
                    distance / 60000
                );

            const seconds =
                Math.floor(
                    (distance % 60000) / 1000
                );

            timerElement.innerText =
                String(minutes).padStart(
                    2, "0"
                )
                + ":"
                +
                String(seconds).padStart(
                    2, "0"
                );

        }}, 1000);

    </script>

    </body>
    </html>
    """

    return render_template_string(html)


# =========================================================
# STUDENT LOGOUT
# =========================================================

@app.route("/student/logout")
def student_logout():

    session.pop("student_uid", None)

    return redirect(
        url_for("student_home")
    )


# =========================================================
# ADMIN SCANNER
# =========================================================

@app.route("/admin/reports")
def admin_reports():
    if not admin_required(["MAIN", "BOYS", "GIRLS"]): return redirect(url_for("admin_login"))
    today = current_time().strftime("%Y-%m-%d")
    month = current_time().strftime("%Y-%m")
    students = list_student_records(gender=admin_scope_gender())
    options = "".join(f'<option value="{escape(s["roll_number"])}">{escape(s["name"])} - {escape(s["roll_number"])}</option>' for s in students)
    return f"""<!DOCTYPE html><html><head><title>SmartMess Reports</title>{CSS}</head><body><div class="container">
    <div class="nav"><a class="btn" href="/admin/dashboard">Dashboard</a><a class="btn gray" href="/admin/records">Records</a></div>
    <div class="card"><h1>📥 Excel & PDF Reports</h1><p>Today, monthly, meal-wise, gender-wise and individual student reports.</p>
    <form action="/admin/report/download" method="get">
      <div class="grid">
       <div><label>From Date</label><input type="date" name="start" value="{today}"></div>
       <div><label>To Date</label><input type="date" name="end" value="{today}"></div>
       <div><label>Gender</label><select name="gender"><option value="">All</option><option value="BOY">Boys</option><option value="GIRL">Girls</option></select></div>
       <div><label>Meal</label><select name="meal"><option value="">All Meals</option><option>BREAKFAST</option><option>LUNCH</option><option>DINNER</option></select></div>
       <div><label>Individual Student</label><select name="registration"><option value="">All Students</option>{options}</select></div>
       <div><label>Format</label><select name="format"><option value="xlsx">Excel (.xlsx)</option><option value="pdf">PDF</option></select></div>
      </div><button class="btn green" type="submit">Download Report</button>
      <a class="btn blue" href="/admin/report/download?start={month}-01&end={today}&format=xlsx">This Month Excel</a>
    </form></div></div></body></html>"""


@app.route("/admin/report/download")
def admin_report_download():
    if not admin_required(["MAIN", "BOYS", "GIRLS"]): return redirect(url_for("admin_login"))
    today = current_time().strftime("%Y-%m-%d")
    start, end = request.args.get("start", today), request.args.get("end", today)
    gender = admin_scope_gender() or request.args.get("gender", "")
    meal = request.args.get("meal", "")
    registration = request.args.get("registration", "")
    records = list_used_records(start, end, gender, meal, registration)
    title = f"SmartMess Attendance Report ({start} to {end})"
    headers = ["Date", "Time", "Name", "Registration No.", "Gender", "Branch", "Hostel", "Block", "Room", "Meal"]
    rows = []
    for r in records:
        used = r.get("used_at") or r.get("generated_at") or ""
        rows.append([used[:10], used[11:19], r.get("name", "-"), r.get("roll_number", "-"), r.get("gender", "-"),
                     r.get("branch", "-"), r.get("hostel_name", "-"), r.get("hostel_block", "-"), r.get("hostel_room", "-"), r.get("meal", "-")])
    fmt = request.args.get("format", "xlsx")
    if fmt == "pdf":
        buffer = BytesIO(); doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm, topMargin=12*mm, bottomMargin=12*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("SmartMessTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22, textColor=colors.HexColor("#102b4f"), spaceAfter=8)
        college_style = ParagraphStyle("CollegeTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=20, leading=24, textColor=colors.HexColor("#102b4f"), alignment=1)
        body_style = ParagraphStyle("SmartMessBody", parent=styles["Normal"], fontName="Helvetica", fontSize=9, leading=12)
        if os.path.exists(COLLEGE_LOGO):
            logo = RLImage(COLLEGE_LOGO, width=22*mm, height=22*mm)
            college_header = Table([[logo, Paragraph(COLLEGE_NAME, college_style)]], colWidths=[28*mm, 235*mm])
            college_header.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),("ALIGN",(1,0),(1,0),"CENTER"),("BOX",(0,0),(-1,-1),.7,colors.HexColor("#dbe3ee")),("BACKGROUND",(0,0),(-1,-1),colors.white),("PADDING",(0,0),(-1,-1),7)]))
            story = [college_header, Spacer(1, 10)]
        else:
            story = [Paragraph(COLLEGE_NAME, college_style), Spacer(1, 8)]
        story += [Paragraph(title, title_style), Paragraph(f"Total attendance records: {len(rows)}", body_style), Spacer(1, 10)]
        table = Table([headers] + rows, repeatRows=1, colWidths=[23*mm,19*mm,33*mm,31*mm,18*mm,29*mm,25*mm,16*mm,16*mm,22*mm])
        table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#102b4f")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),7),("GRID",(0,0),(-1,-1),.35,colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f8fafc")]),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("PADDING",(0,0),(-1,-1),4)]))
        story.append(table); doc.build(story); buffer.seek(0)
        return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=f"SmartMess_Report_{start}_{end}.pdf")
    wb = Workbook(); ws = wb.active; ws.title = "Attendance"
    ws.merge_cells("B1:J1"); ws["B1"] = COLLEGE_NAME; ws["B1"].font = Font(size=20, bold=True, color="102B4F"); ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B2:J2"); ws["B2"] = title; ws["B2"].font = Font(size=15, bold=True, color="FFFFFF"); ws["B2"].fill = PatternFill("solid", fgColor="102B4F"); ws["B2"].alignment = Alignment(horizontal="center")
    ws["A3"] = f"Total records: {len(rows)}"; ws.append([]); ws.append(headers)
    if os.path.exists(COLLEGE_LOGO):
        logo = XLImage(COLLEGE_LOGO); logo.width = 62; logo.height = 62; ws.add_image(logo, "A1")
    ws.row_dimensions[1].height = 48; ws.row_dimensions[2].height = 25
    for cell in ws[5]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="2563EB")
    for row in rows: ws.append(row)
    thin = Side(style="thin", color="D9E2F0")
    for row in ws.iter_rows(min_row=5):
        for cell in row: cell.border = Border(bottom=thin); cell.alignment = Alignment(vertical="center")
    widths = [13,11,24,21,12,25,20,12,12,15]
    for i, width in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A6"; ws.auto_filter.ref = f"A5:J{max(5, ws.max_row)}"
    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return send_file(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"SmartMess_Report_{start}_{end}.xlsx")


@app.route("/admin/menu-notice", methods=["GET", "POST"])
def admin_menu_notice():
    if not admin_required(["MAIN", "BOYS", "GIRLS"]): return redirect(url_for("admin_login"))
    message = ""
    if request.method == "POST":
        save_settings({key: request.form.get(key, "").strip() for key in ["breakfast_menu", "lunch_menu", "dinner_menu", "student_notice"]}); message = "Menu and notice saved."
    values = {key: get_setting(key, "") for key in ["breakfast_menu", "lunch_menu", "dinner_menu", "student_notice"]}
    return f"""<!DOCTYPE html><html><head><title>Menu & Notice</title>{CSS}</head><body><div class="container">
    <div class="nav"><a class="btn" href="/admin/dashboard">Dashboard</a></div><div class="card"><h1>📋 Menu & Notice</h1>
    {f'<div class="message success">{message}</div>' if message else ''}<form method="post">
    <label>Breakfast Menu</label><input name="breakfast_menu" value="{escape(values['breakfast_menu'])}">
    <label>Lunch Menu</label><input name="lunch_menu" value="{escape(values['lunch_menu'])}">
    <label>Dinner Menu</label><input name="dinner_menu" value="{escape(values['dinner_menu'])}">
    <label>Student Notice</label><input name="student_notice" value="{escape(values['student_notice'])}">
    <button class="btn green">Save Menu & Notice</button></form></div></div></body></html>"""


@app.route("/admin/roles", methods=["GET", "POST"])
def admin_roles():
    if not admin_required(["MAIN"]): return redirect(url_for("admin_dashboard"))
    message = ""
    if request.method == "POST":
        username, password, role = request.form.get("username", "").strip(), request.form.get("password", ""), request.form.get("role", "")
        if len(username) < 3 or len(password) < 8 or role not in ["BOYS", "GIRLS", "SCANNER"]: message = "Use 3+ character username and 8+ character password."
        elif add_admin_account(username, password, role): message = "Admin account created successfully."
        else: message = "Username already exists."
    rows = "".join(f"<tr><td>{escape(a['username'])}</td><td>{a['role']}</td><td>{'Active' if a.get('active',1) else 'Inactive'}</td><td>{escape(a.get('last_login') or 'Never')}</td><td><form method='post' action='/admin/role/{a['id']}/toggle'><button class='btn gray'>Toggle</button></form><form method='post' action='/admin/role/{a['id']}/reset'><input name='password' type='password' minlength='8' placeholder='New password' required><button class='btn blue'>Reset</button></form></td></tr>" for a in list_admins())
    return f"""<!DOCTYPE html><html><head><title>Admin Roles</title>{CSS}</head><body><div class="container"><div class="nav"><a class="btn" href="/admin/dashboard">Dashboard</a></div>
    <div class="card"><h1>🔐 Admin Roles</h1>{f'<div class="message">{escape(message)}</div>' if message else ''}<form method="post"><div class="grid">
    <div><label>Username</label><input name="username" required></div><div><label>Password</label><input type="password" name="password" minlength="8" required></div>
    <div><label>Role</label><select name="role"><option value="BOYS">Boys Hostel Admin</option><option value="GIRLS">Girls Hostel Admin</option><option value="SCANNER">Scanner Operator</option></select></div></div>
    <button class="btn green">Create Account</button></form></div><div class="card"><table><tr><th>Username</th><th>Role</th><th>Status</th><th>Last Login</th><th>Actions</th></tr>{rows}</table></div></div></body></html>"""


@app.route("/admin/role/<int:admin_id>/toggle",methods=["POST"])
def admin_role_toggle(admin_id):
    if not admin_required(["MAIN"]):return redirect(url_for("admin_login"))
    account=next((a for a in list_admins() if a["id"]==admin_id),None)
    if account:
        active=0 if account.get("active",1) else 1
        if USE_MONGO:mongo_database.admins.update_one({"id":admin_id},{"$set":{"active":active}})
        else:
            conn=db();conn.execute("UPDATE admins SET active=? WHERE id=?",(active,admin_id));conn.commit();conn.close()
        log_activity("ADMIN_STATUS_CHANGED",f"{account['username']} active={active}")
    return redirect(url_for("admin_roles"))


@app.route("/admin/role/<int:admin_id>/reset",methods=["POST"])
def admin_role_reset(admin_id):
    if not admin_required(["MAIN"]):return redirect(url_for("admin_login"))
    password=request.form.get("password","")
    if len(password)>=8:
        account=next((a for a in list_admins() if a["id"]==admin_id),None); now=current_time().strftime("%Y-%m-%d %H:%M:%S"); password_hash=generate_password_hash(password)
        if USE_MONGO:mongo_database.admins.update_one({"id":admin_id},{"$set":{"password_hash":password_hash,"password_changed_at":now}})
        else:
            conn=db();conn.execute("UPDATE admins SET password_hash=?,password_changed_at=? WHERE id=?",(password_hash,now,admin_id));conn.commit();conn.close()
        if account:log_activity("ADMIN_PASSWORD_RESET",account["username"])
    return redirect(url_for("admin_roles"))

# =========================================================
# ADMIN QR SCANNER
# =========================================================

@app.route("/admin/scanner")
def admin_scanner():

    if not admin_required():
        return redirect(url_for("admin_login"))

    html = f"""
    <!DOCTYPE html>
    <html>

    <head>

        <title>Mess QR Scanner</title>

        {CSS}

        <script src="/static/html5-qrcode.min.js?v=2.3.8"></script>

        <style>
            #reader {{
                width: 100%;
                max-width: 600px;
                margin: 20px auto;
            }}

            #result {{
                margin-top: 20px;
            }}

            .big-photo {{
                width: 230px;
                height: 230px;
                object-fit: cover;
                border-radius: 15px;
                display: block;
                margin: 15px auto;
            }}
        </style>

    </head>

    <body>

    <div class="container">

        <div class="nav">

            <a class="btn"
               href="/admin/dashboard">
                Dashboard
            </a>

            <a class="btn gray"
               href="/admin/records">
                📊 Records
            </a>

            <a class="btn red"
               href="/admin/logout">
                Logout
            </a>

        </div>

        <div class="card">

            <h1>📷 Mess QR Scanner</h1>

            <p>
                Allow camera permission and scan the student's coupon QR.
            </p>

            <div id="reader"></div>

            <div id="result"></div>

        </div>

    </div>

    <script>

    let processing = false;

    function sound(ok) {{
        const audio = new (window.AudioContext || window.webkitAudioContext)();
        const oscillator = audio.createOscillator(); const gain = audio.createGain();
        oscillator.connect(gain); gain.connect(audio.destination);
        oscillator.frequency.value = ok ? 880 : 190; gain.gain.value = .22;
        oscillator.start(); oscillator.stop(audio.currentTime + (ok ? .22 : .55));
    }}

    async function verifyCoupon(token) {{

        if (processing) {{
            return;
        }}

        processing = true;

        try {{

            const response = await fetch(
                "/admin/verify-coupon",
                {{
                    method: "POST",
                    headers: {{
                        "Content-Type": "application/json"
                    }},
                    body: JSON.stringify({{
                        token: token
                    }})
                }}
            );

            const data = await response.json();

            if (data.success) {{

                sound(true);

                const photo =
                    data.photo
                    ?
                    "<img class='big-photo' src='"
                    + data.photo
                    + "'>"
                    :
                    "";

                document.getElementById(
                    "result"
                ).innerHTML =

                    "<div class='message success'>"

                    + "<h1 style='font-size:42px'>✅ MEAL APPROVED</h1>"

                    + photo

                    + "<h2>"
                    + data.name
                    + "</h2>"

                    + "<p>Registration No.: "
                    + data.roll
                    + "</p>"

                    + "<p>Branch: "
                    + data.branch
                    + "</p>"

                    + "<p>Room: "
                    + data.room
                    + "</p>"

                    + "<p>Gender: " + (data.gender || "NOT SET") + "</p>"
                    + "<p>Hostel: " + (data.hostel || "NOT SET") + " " + (data.block || "") + "</p>"

                    + "<p>Meal: <b>"
                    + data.meal
                    + "</b></p>"

                    + "<p>Coupon is now USED.</p>"

                    + "</div>";

            }} else {{

                sound(false);
                const alreadyUsed = (data.message || '').toLowerCase().includes('already used');

                document.getElementById(
                    "result"
                ).innerHTML =

                    "<div class='message error'>"

                    + "<h1 style='font-size:42px'>❌ " + (alreadyUsed ? "ALREADY USED" : "COUPON REJECTED") + "</h1>"

                    + "<p>"
                    + data.message
                    + "</p>"

                    + (data.can_extend ? "<button class='btn blue' id='extendCouponBtn'>Add 5 Minutes</button>" : "")

                    + "</div>";

                if (data.can_extend) {{
                    document.getElementById("extendCouponBtn").onclick = function() {{
                        extendCoupon(data.token);
                    }};
                }}
            }}

        }} catch (error) {{

            document.getElementById(
                "result"
            ).innerHTML =

                "<div class='message error'>"

                + "<h2>❌ Scanner Error</h2>"

                + "<p>"
                + error
                + "</p>"

                + "</div>";
        }}

        setTimeout(
            function() {{
                processing = false;
            }},
            2000
        );

    }}

    async function extendCoupon(token) {{
      const response=await fetch('/admin/extend-coupon',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{token}})}});
      const data=await response.json();
      document.getElementById('result').innerHTML='<div class="message '+(data.success?'success':'error')+'"><h2>'+(data.success?'⏱️ QR EXTENDED':'❌ NOT EXTENDED')+'</h2><p>'+data.message+'</p></div>';
      sound(data.success); processing=false;
    }}

    function scanSuccess(
        decodedText,
        decodedResult
    ) {{

        verifyCoupon(decodedText);

    }}

    function scanFailure(error) {{

        // Ignore scan errors

    }}

    function startQrScanner() {{
        if (typeof Html5QrcodeScanner === "undefined") return false;
        const scanner = new Html5QrcodeScanner(
            "reader",
            {{
                fps: 10,
                qrbox: {{ width: 250, height: 250 }},
                rememberLastUsedCamera: true,
                supportedScanTypes: [
                    Html5QrcodeScanType.SCAN_TYPE_CAMERA,
                    Html5QrcodeScanType.SCAN_TYPE_FILE
                ]
            }},
            false
        );
        scanner.render(scanSuccess, scanFailure);
        return true;
    }}

    if (!startQrScanner()) {{
        const backup = document.createElement("script");
        backup.src = "https://cdnjs.cloudflare.com/ajax/libs/html5-qrcode/2.3.8/html5-qrcode.min.js";
        backup.onload = function() {{ startQrScanner(); }};
        backup.onerror = function() {{
            document.getElementById("reader").innerHTML =
                '<div class="message error"><h2>Scanner could not load</h2>' +
                '<p>Please check internet connection, disable web protection for this site, and refresh.</p></div>';
        }};
        document.head.appendChild(backup);
    }}

    </script>

    </body>
    </html>
    """

    return html

# =========================================================
# ADMIN RECORDS
# =========================================================

@app.route("/admin/records")
def admin_records():

    if not admin_required():
        return redirect(
            url_for("admin_login")
        )

    records = all_records()
    if admin_scope_gender():
        allowed = {s["student_uid"] for s in list_student_records(gender=admin_scope_gender())}
        records = [r for r in records if r.get("student_uid") in allowed]

    rows = ""

    for record in records:

        if record["status"] == "USED":

            badge = (
                '<span class="badge used-badge">'
                'USED</span>'
            )

        elif record["status"] == "EXPIRED":

            badge = (
                '<span class="badge expired-badge">'
                'EXPIRED</span>'
            )

        else:

            badge = (
                '<span class="badge active-badge">'
                'ACTIVE</span>'
            )

        rows += f"""
        <tr>

            <td>{record["name"] or "-"}</td>

            <td>{record["roll_number"] or "-"}</td>

            <td>{record["meal"]}</td>

            <td>{record["generated_at"]}</td>

            <td>{record["expires_at"]}</td>

            <td>{record["used_at"] or "-"}</td>

            <td>{badge}</td>

        </tr>
        """

    html = f"""
    <!DOCTYPE html>
    <html>

    <head>

        <title>Records</title>

        {CSS}

    </head>

    <body>

    <div class="container">

        <div class="nav">

            <a
                class="btn"
                href="/admin/dashboard"
            >
                Dashboard
            </a>

        </div>

        <div class="card">

            <h1>📊 Mess Records</h1>

            <div style="overflow-x:auto;">

                <table>

                    <tr>
                        <th>Name</th>
                        <th>Registration No.</th>
                        <th>Meal</th>
                        <th>Generated</th>
                        <th>Expires</th>
                        <th>Used At</th>
                        <th>Status</th>
                    </tr>

                    {rows}

                </table>

            </div>

        </div>

    </div>

    </body>

    </html>
    """

    return html


# =========================================================
# PHASE 4: OPERATIONS, SECURITY AND PWA
# =========================================================

@app.route("/admin/weekly-menu",methods=["GET","POST"])
def admin_weekly_menu():
    if not admin_required(["MAIN","BOYS","GIRLS"]):return redirect(url_for("admin_login"))
    menu=weekly_menu(); message=""
    if request.method=="POST":
        for day in menu:
            for meal in menu[day]: menu[day][meal]=request.form.get(f"{day}_{meal}","").strip()[:150]
        save_settings({"weekly_menu":json.dumps(menu)}); message="Weekly menu saved."; log_activity("WEEKLY_MENU_UPDATED")
    rows="".join(f'<tr><td><b>{day}</b></td>'+''.join(f'<td><input name="{day}_{meal}" value="{escape(menu[day][meal])}"></td>' for meal in ["BREAKFAST","LUNCH","DINNER"])+"</tr>" for day in menu)
    return f'''<!doctype html><html><head><title>Weekly Menu</title>{CSS}</head><body><div class="container"><div class="nav"><a class="btn" href="/admin/dashboard">Dashboard</a></div><div class="card"><h1>📅 Weekly Menu</h1>{f'<div class="message success">{message}</div>' if message else ''}<form method="post"><div style="overflow:auto"><table><tr><th>Day</th><th>Breakfast</th><th>Lunch</th><th>Dinner</th></tr>{rows}</table></div><button class="btn green">Save Weekly Menu</button></form></div></div></body></html>'''


@app.route("/admin/complaints",methods=["GET","POST"])
def admin_complaints():
    if not admin_required(["MAIN","BOYS","GIRLS"]):return redirect(url_for("admin_login"))
    if request.method=="POST":
        cid=int(request.form.get("id","0")); status=request.form.get("status",""); reply=request.form.get("reply","").strip()[:500]
        if status in ["OPEN","IN_PROGRESS","RESOLVED"]:
            changes={"status":status,"admin_reply":reply,"updated_at":current_time().strftime("%Y-%m-%d %H:%M:%S")}
            if USE_MONGO:mongo_database.complaints.update_one({"id":cid},{"$set":changes})
            else:
                conn=db(); conn.execute("UPDATE complaints SET status=?,admin_reply=?,updated_at=? WHERE id=?",(status,reply,changes["updated_at"],cid));conn.commit();conn.close()
            log_activity("COMPLAINT_UPDATED",f"Complaint #{cid}: {status}")
    students={s["student_uid"]:s for s in list_student_records(gender=admin_scope_gender())}
    cards=""
    for c in list_complaints():
        s=students.get(c["student_uid"])
        if not s:continue
        cards+=f'''<div class="card"><h3>{escape(s['name'])} · {escape(c['category'])} · {'★'*int(c.get('rating',0))}</h3><p>{escape(c['message'])}</p><form method="post"><input type="hidden" name="id" value="{c['id']}"><select name="status"><option {'selected' if c.get('status')=='OPEN' else ''}>OPEN</option><option {'selected' if c.get('status')=='IN_PROGRESS' else ''}>IN_PROGRESS</option><option {'selected' if c.get('status')=='RESOLVED' else ''}>RESOLVED</option></select><input name="reply" value="{escape(c.get('admin_reply','') or '')}" placeholder="Admin reply"><button class="btn blue">Update</button></form></div>'''
    return f'''<!doctype html><html><head><title>Complaints</title>{CSS}</head><body><div class="container"><div class="nav"><a class="btn" href="/admin/dashboard">Dashboard</a></div><h1>💬 Complaints & Feedback</h1>{cards or '<div class="card">No complaints.</div>'}</div></body></html>'''


@app.route("/admin/pin-requests")
def admin_pin_requests():
    if not admin_required(["MAIN","BOYS","GIRLS"]):return redirect(url_for("admin_login"))
    allowed={s["student_uid"]:s for s in list_student_records(gender=admin_scope_gender())}; rows=""
    for item in list_pin_requests():
        s=allowed.get(item["student_uid"])
        if not s:continue
        action=f'<form method="post" action="/admin/pin-request/{item["id"]}/reset"><button class="btn green">Generate Temporary PIN</button></form>' if item["status"]=="PENDING" else "Resolved"
        rows+=f'<tr><td>{escape(s["name"])}</td><td>{escape(s["roll_number"])}</td><td>{escape(item.get("hostel_room", ""))}</td><td>{escape(item.get("reason", ""))}</td><td>{item["status"]}</td><td>{action}</td></tr>'
    return f'''<!doctype html><html><head><title>PIN Requests</title>{CSS}</head><body><div class="container"><div class="nav"><a class="btn" href="/admin/dashboard">Dashboard</a></div><div class="card"><h1>🔑 Forgot PIN Requests</h1><div style="overflow:auto"><table><tr><th>Student</th><th>Registration</th><th>Room</th><th>Reason</th><th>Status</th><th>Action</th></tr>{rows or '<tr><td colspan="6">No requests.</td></tr>'}</table></div></div></div></body></html>'''


@app.route("/admin/pin-request/<int:request_id>/reset",methods=["POST"])
def admin_pin_reset(request_id):
    if not admin_required(["MAIN","BOYS","GIRLS"]):return redirect(url_for("admin_login"))
    items=[r for r in list_pin_requests() if r["id"]==request_id]
    if not items:return "Request not found",404
    item=items[0]; student=student_by_uid(item["student_uid"])
    if not student or (admin_scope_gender() and student.get("gender")!=admin_scope_gender()):return "Not allowed",403
    pin=f"{secrets.randbelow(10000):04d}"; update_student_record(student["id"],{"pin_hash":generate_password_hash(pin),"force_pin_change":1,"pin_changed_at":current_time().strftime("%Y-%m-%d %H:%M:%S")})
    changes=("RESOLVED",current_time().strftime("%Y-%m-%d %H:%M:%S"),session.get("admin_username"))
    if USE_MONGO:mongo_database.pin_requests.update_one({"id":request_id},{"$set":{"status":changes[0],"resolved_at":changes[1],"resolved_by":changes[2]}})
    else:
        conn=db();conn.execute("UPDATE pin_requests SET status=?,resolved_at=?,resolved_by=? WHERE id=?",(*changes,request_id));conn.commit();conn.close()
    log_activity("STUDENT_PIN_RESET",f"Student {student['roll_number']}")
    return f'''<!doctype html><html><head>{CSS}</head><body><div class="container"><div class="card center"><h1>Temporary PIN</h1><p>Give this PIN privately to <b>{escape(student['name'])}</b>. It is shown only on this screen.</p><div class="countdown">{pin}</div><p>Student must change it after login.</p><a class="btn" href="/admin/pin-requests">Done</a></div></div></body></html>'''


@app.route("/admin/activity-log")
def admin_activity_log():
    if not admin_required(["MAIN"]):return redirect(url_for("admin_dashboard"))
    rows="".join(f'<tr><td>{escape(x["created_at"])}</td><td>{escape(x.get("actor", ""))}</td><td>{escape(x.get("role", ""))}</td><td>{escape(x["action"])}</td><td>{escape(x.get("details", ""))}</td></tr>' for x in list_activity())
    return f'''<!doctype html><html><head><title>Activity Log</title>{CSS}</head><body><div class="container"><div class="nav"><a class="btn" href="/admin/dashboard">Dashboard</a></div><div class="card"><h1>🛡️ Admin Activity Log</h1><div style="overflow:auto"><table><tr><th>Time</th><th>Admin</th><th>Role</th><th>Action</th><th>Details</th></tr>{rows or '<tr><td colspan="5">No activity.</td></tr>'}</table></div></div></div></body></html>'''


@app.route("/admin/extend-coupon",methods=["POST"])
def extend_coupon():
    if not admin_required():return jsonify({"success":False,"message":"Admin login required."}),401
    token=str((request.get_json(silent=True) or {}).get("token","")).strip(); coupon=coupon_by_token(token)
    if not coupon or coupon.get("status")=="USED":return jsonify({"success":False,"message":"Coupon cannot be extended."})
    if (coupon.get("generated_at") or "")[:10]!=current_time().date().isoformat() or int(coupon.get("extension_count",0) or 0)>=1:return jsonify({"success":False,"message":"Only today's expired coupon can be extended once."})
    expiry=datetime.strptime(coupon["expires_at"],"%Y-%m-%d %H:%M:%S")
    if current_time()<=expiry:return jsonify({"success":False,"message":"Coupon is still active."})
    new_expiry=(current_time()+timedelta(minutes=5)).strftime("%Y-%m-%d %H:%M:%S")
    update_coupon(coupon["id"],{"expires_at":new_expiry,"status":"ACTIVE","extension_count":1,"extended_at":current_time().strftime("%Y-%m-%d %H:%M:%S"),"extended_by":session.get("admin_username")})
    log_activity("QR_EXTENDED",f"Coupon {token[-6:]} +5 minutes")
    return jsonify({"success":True,"message":"Coupon extended for 5 minutes. Scan it again."})


@app.route("/manifest.json")
def manifest():
    return jsonify({"name":"SmartMess","short_name":"SmartMess","start_url":"/student","display":"standalone","background_color":"#f4f7fb","theme_color":"#081b33","icons":[{"src":"/static/icon-192.png","sizes":"192x192","type":"image/png"},{"src":"/static/icon-512.png","sizes":"512x512","type":"image/png"}]})


@app.route("/service-worker.js")
def service_worker():
    script="""const C='smartmess-v4';self.addEventListener('install',e=>e.waitUntil(caches.open(C).then(c=>c.addAll(['/offline','/static/icon-192.png','/static/icon-512.png']))));self.addEventListener('fetch',e=>{if(e.request.mode==='navigate'){e.respondWith(fetch(e.request).catch(()=>caches.match('/offline')))}});"""
    return Response(script,mimetype="application/javascript",headers={"Service-Worker-Allowed":"/"})


@app.route("/offline")
def offline():
    return f'''<!doctype html><html><head>{CSS}</head><body><div class="container"><div class="card center"><h1>📴 You are offline</h1><p>Login, coupon generation and scanning need an internet connection.</p></div></div></body></html>'''


# =========================================================
# RUN
# =========================================================

if __name__ == "__main__":

    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )
